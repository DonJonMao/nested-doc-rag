from __future__ import annotations

import base64
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from nested_doc_rag.excel.comments import format_source_comment
from nested_doc_rag.excel.writeback import patch_workbook, prepare_writeback_item
from nested_doc_rag.io import read_json, read_jsonl, write_jsonl
from nested_doc_rag.schemas.eval import FieldPrediction

TINY_PNG = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
)


def make_prediction(
    field_id: str,
    target_cell: str,
    value: object,
    *,
    status: str = "answered",
    confidence: float = 0.91,
    validation: dict | None = None,
    reference_docs: list[dict] | None = None,
    evidence_ids: list[str] | None = None,
) -> FieldPrediction:
    return FieldPrediction(
        field_id=field_id,
        row_index=4,
        target_cell=target_cell,
        answer_value=value,
        answer_status=status,
        confidence=confidence,
        source_chunk_ids=["chunk_1"],
        evidence_attachment_ids=["img_1"] if evidence_ids is None else evidence_ids,
        reference_source_documents=reference_docs or [],
        reference_chunk_ids=[str(doc.get("chunk_id")) for doc in reference_docs or [] if doc.get("chunk_id")],
        validation=validation or {},
    )


def make_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["C4"] = "old answer"
    worksheet["C4"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    worksheet["C4"].font = Font(bold=True, color="FF0000")
    worksheet["C4"].alignment = Alignment(horizontal="center")
    worksheet.column_dimensions["C"].width = 28
    worksheet.row_dimensions[4].height = 30
    worksheet["C5"] = "keep me"
    worksheet["C6"] = None
    worksheet["D4"] = "=SUM(A1:A2)"
    worksheet["A1"] = 1
    worksheet["A2"] = 2
    worksheet.merge_cells("E1:F1")
    worksheet["E1"] = "merged"
    workbook.save(path)


def write_tiny_png(path: Path) -> None:
    path.write_bytes(base64.b64decode(TINY_PNG))


def assert_adjacent_evidence_layout(workbook, *, evidence_cell: str = "D6", image_count: int = 0) -> None:
    assert "Evidence" not in workbook.sheetnames
    assert workbook["Sheet1"][evidence_cell].value
    assert len(workbook["Sheet1"]._images) == image_count


def test_prepare_writeback_item() -> None:
    item = prepare_writeback_item("Sheet1", "A1", "未找到", comment="source missing")
    assert item.sheet_name == "Sheet1"
    assert item.cell == "A1"
    assert item.value == "未找到"
    assert item.comment == "source missing"


def test_format_source_comment() -> None:
    assert format_source_comment("chunk_1") == "chunk_1"
    assert format_source_comment("chunk_1", "low confidence") == "chunk_1\nlow confidence"


def test_write_answered_cell(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    summary = patch_workbook(template, [make_prediction("field_1", "C4", "new answer")], output)

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C4"].value == "new answer"
    assert summary.written_count == 1
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["action"] == "written"
    assert audit[0]["reason"] == "written"


def test_skip_not_found(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    summary = patch_workbook(template, [make_prediction("field_1", "C5", "未找到", status="not_found")], output)

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C5"].value == "keep me"
    assert summary.written_count == 0
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    review_items = read_jsonl(tmp_path / "review_items.jsonl")
    assert audit[0]["reason"] == "skipped_status"
    assert review_items[0]["reason"] == "skipped_status"


def test_preserve_style(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    patch_workbook(template, [make_prediction("field_1", "C4", "new answer")], output)

    workbook = load_workbook(output)
    worksheet = workbook["Sheet1"]
    cell = worksheet["C4"]
    assert cell.value == "new answer"
    assert cell.fill.fgColor.rgb == "00FFFF00"
    assert cell.font.bold is True
    assert cell.font.color.rgb == "00FF0000"
    assert cell.alignment.horizontal == "center"
    assert worksheet.column_dimensions["C"].width == 28
    assert worksheet.row_dimensions[4].height >= 30
    assert "E1:F1" in {str(item) for item in worksheet.merged_cells.ranges}


def test_comment_contains_evidence(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    patch_workbook(
        template,
        [
            make_prediction(
                "field_1",
                "C4",
                "new answer",
                confidence=0.91234,
                reference_docs=[
                    {
                        "chunk_id": "chunk_1",
                        "file_name": "能力清单.xlsx",
                        "text_preview": "机房供电采用双路市电，配备 UPS。",
                    }
                ],
            )
        ],
        output,
        trace_by_field={"field_1": "trace_001"},
    )

    workbook = load_workbook(output)
    comment = workbook["Sheet1"]["C4"].comment
    assert comment is not None
    assert "document: 能力清单.xlsx" in comment.text
    assert "text: 机房供电采用双路市电，配备 UPS。" in comment.text
    assert "source_chunk_ids" not in comment.text
    assert "evidence_attachment_ids" not in comment.text
    assert "trace_id" not in comment.text
    evidence_map = read_json(tmp_path / "evidence_map.json")
    assert evidence_map["fields"]["field_1"]["source_chunk_ids"] == ["chunk_1"]


def test_skip_formula_cell(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    summary = patch_workbook(template, [make_prediction("field_1", "D4", "3")], output)

    workbook = load_workbook(output, data_only=False)
    assert workbook["Sheet1"]["D4"].value == "=SUM(A1:A2)"
    assert summary.formula_skipped_count == 1
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["reason"] == "skipped_formula"


def test_duplicate_target_cell_conflict(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    summary = patch_workbook(
        template,
        [
            make_prediction("field_1", "C4", "first"),
            make_prediction("field_2", "C4", "second"),
        ],
        output,
    )

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C4"].value == "old answer"
    assert summary.conflict_count == 2
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert {record["action"] for record in audit} == {"conflict"}
    assert {record["reason"] for record in audit} == {"duplicate_target_cell"}


def test_invalid_target_cell_audit(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    summary = patch_workbook(template, [make_prediction("field_1", "Missing!C4", "new answer")], output)

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C4"].value == "old answer"
    assert summary.invalid_count == 1
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    review_items = read_jsonl(tmp_path / "review_items.jsonl")
    assert audit[0]["action"] == "invalid"
    assert audit[0]["reason"] == "invalid_cell"
    assert review_items[0]["reason"] == "invalid_cell"


def test_uncertain_default_off_goes_to_review_only(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "document_id": "doc_1",
                "object_key": "kb/xixian/doc.xlsx",
                "source_anchor": "能力清单!H42",
                "sheet_name": "能力清单",
                "cell": "H42",
                "text_preview": "市电配置为双路。",
            }
        ],
    )

    summary = patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
    )

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C6"].value is None
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["status"] == "uncertain"
    assert audit[0]["writeback_action"] == "skipped_uncertain_policy"
    assert audit[0]["error_code"] == "WB_POLICY_REJECTED"
    assert summary.uncertain_count == 1
    assert summary.written_count == 0


def test_uncertain_allowed_writes_red_comment_and_evidence(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "document_id": "doc_1",
                "object_key": "kb/xixian/doc.xlsx",
                "source_anchor": "能力清单!H42",
                "sheet_name": "能力清单",
                "cell": "H42",
                "text_preview": "市电配置为双路。",
            }
        ],
        evidence_ids=[],
    )

    summary = patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True, "uncertain_comment_prefix": "[UNCERTAIN]"},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    cell = workbook["Sheet1"]["C6"]
    assert cell.value == "双路市电"
    assert cell.fill.fgColor.rgb == "FFFFCCCC"
    assert cell.comment is not None
    assert "[UNCERTAIN]" in cell.comment.text
    assert "document: doc_1" in cell.comment.text
    assert "市电配置为双路" in cell.comment.text
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    review_items = read_jsonl(tmp_path / "review_items.jsonl")
    assert audit[0]["status"] == "uncertain"
    assert audit[0]["writeback_action"] == "written_red_comment"
    assert audit[0]["evidence_count"] == 1
    assert review_items[0]["status"] == "uncertain"
    assert summary.uncertain_count == 1
    assert summary.written_count == 1
    assert summary.review_count == 1
    assert_adjacent_evidence_layout(workbook)
    assert "市电配置为双路" in workbook["Sheet1"]["D6"].value


def test_uncertain_comment_is_truncated_to_configured_limit(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)
    limit = 120

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "document_id": "doc_1",
                "object_key": "kb/xixian/doc.xlsx",
                "source_anchor": "能力清单!H42",
                "sheet_name": "能力清单",
                "cell": "H42",
                "text_preview": "市电配置为双路。" * 80,
            }
        ],
        evidence_ids=[],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True, "max_comment_chars": limit},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    comment = workbook["Sheet1"]["C6"].comment
    assert comment is not None
    assert len(comment.text) <= limit
    assert "WB_COMMENT_TOO_LONG" in comment.text
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["comment_length"] <= limit
    assert audit[0]["error_code"] == "WB_COMMENT_TOO_LONG"


def test_flagged_does_not_write(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    patch_workbook(
        template,
        [make_prediction("field_1", "C5", "未找到", status="not_found")],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
    )

    workbook = load_workbook(output)
    assert workbook["Sheet1"]["C5"].value == "keep me"
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["status"] == "flagged"
    assert audit[0]["writeback_action"] == "review_only"


def test_proof_attachment_ids_generate_image_evidence_artifact(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "document_id": "doc_1",
                "source_anchor": "能力清单!H42",
                "proof_attachment_ids": ["proof_img_1"],
            }
        ],
        evidence_ids=[],
    )

    summary = patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    assert summary.image_evidence_path == tmp_path / "image_evidence.jsonl"
    image_rows = read_jsonl(tmp_path / "image_evidence.jsonl")
    assert image_rows[0]["field_id"] == "field_1"
    assert image_rows[0]["image_object_key"] == "runs/run_1/evidence/field_1/proof_img_1"
    audit = read_jsonl(tmp_path / "writeback_audit.jsonl")
    assert audit[0]["image_evidence_count"] == 1
    assert audit[0]["evidence_refs"][0]["image_object_key"] == "runs/run_1/evidence/field_1/proof_img_1"
    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook)
    assert "runs/run_1/evidence/field_1/proof_img_1" in workbook["Sheet1"]["D6"].value
    assert "图片不可用" in workbook["Sheet1"]["E6"].value


def test_adjacent_columns_append_local_image_proof(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    image_path = tmp_path / "proof.png"
    make_template(template)
    write_tiny_png(image_path)

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "file_name": "能力清单.xlsx",
                "text_preview": "供电采用双路市电。",
                "proof_attachment_ids": ["proof_img_1"],
                "proof_attachments": [
                    {
                        "attachment_id": "proof_img_1",
                        "image_path": str(image_path),
                        "media_content_type": "image/png",
                    }
                ],
            }
        ],
        evidence_ids=[],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook, image_count=1)
    assert "供电采用双路市电" in workbook["Sheet1"]["D6"].value


def test_adjacent_columns_extracts_xlsx_media_image(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    source_workbook = tmp_path / "source.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)
    with zipfile.ZipFile(source_workbook, "w") as archive:
        archive.writestr("xl/media/proof.png", base64.b64decode(TINY_PNG))

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "file_name": "source.xlsx",
                "relative_path": "source.xlsx",
                "text_preview": "供电采用双路市电。",
                "proof_attachment_ids": ["proof_img_1"],
                "proof_attachments": [
                    {
                        "attachment_id": "proof_img_1",
                        "media_path": "xl/media/proof.png",
                        "media_content_type": "image/png",
                    }
                ],
            }
        ],
        evidence_ids=[],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook, image_count=1)
    assert "source.xlsx" in workbook["Sheet1"]["D6"].value


def test_adjacent_columns_resolves_dispimg_media_from_attachment_id(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    source_workbook = tmp_path / "source.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    source = Workbook()
    source_sheet = source.active
    source_sheet.title = "Source"
    source_sheet["E3"] = '=_xlfn.DISPIMG("ID_TEST_IMAGE",1)'
    source.save(source_workbook)
    with zipfile.ZipFile(source_workbook, "a") as archive:
        archive.writestr("xl/media/proof.png", base64.b64decode(TINY_PNG))
        archive.writestr(
            "xl/cellimages.xml",
            """
            <etc:cellImages xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
              xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
              xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
              xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">
              <etc:cellImage>
                <xdr:pic>
                  <xdr:nvPicPr><xdr:cNvPr id="1" name="ID_TEST_IMAGE"/></xdr:nvPicPr>
                  <xdr:blipFill><a:blip r:embed="rId1"/></xdr:blipFill>
                </xdr:pic>
              </etc:cellImage>
            </etc:cellImages>
            """,
        )
        archive.writestr(
            "xl/_rels/cellimages.xml.rels",
            """
            <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
              <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/proof.png"/>
            </Relationships>
            """,
        )

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[
            {
                "chunk_id": "chunk_1",
                "file_name": "source.xlsx",
                "relative_path": "source.xlsx",
                "sheet_name": "Source",
                "cell": "A3:L3",
                "text_preview": "供电采用双路市电。",
                "proof_attachment_ids": ["att_file_abc_E3_dispimg"],
            }
        ],
        evidence_ids=[],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook, image_count=1)
    assert "A3:L3" in workbook["Sheet1"]["D6"].value


def test_adjacent_columns_uses_proof_attachment_registry_for_attachment_only_prediction(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled_form.xlsx"
    image_path = tmp_path / "registry-proof.png"
    make_template(template)
    write_tiny_png(image_path)
    write_jsonl(
        tmp_path / "proof_attachment_registry.jsonl",
        [
            {
                "attachment_id": "att_registry_img",
                "image_path": str(image_path),
                "file_name": "能力清单.xlsx",
                "sheet_name": "能力清单",
                "source_cell": "E3",
            }
        ],
    )

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[],
        evidence_ids=["att_registry_img"],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook, image_count=1)
    assert "att_registry_img" in workbook["Sheet1"]["D6"].value


def test_adjacent_columns_resolves_attachment_only_dispimg_from_manifest(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    source_workbook = tmp_path / "source.xlsx"
    output = tmp_path / "filled_form.xlsx"
    make_template(template)

    source = Workbook()
    source_sheet = source.active
    source_sheet.title = "Source"
    source_sheet["E3"] = '=_xlfn.DISPIMG("ID_TEST_IMAGE",1)'
    source.save(source_workbook)
    with zipfile.ZipFile(source_workbook, "a") as archive:
        archive.writestr("xl/media/proof.png", base64.b64decode(TINY_PNG))
        archive.writestr(
            "xl/cellimages.xml",
            """
            <etc:cellImages xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
              xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
              xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
              xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">
              <etc:cellImage>
                <xdr:pic>
                  <xdr:nvPicPr><xdr:cNvPr id="1" name="ID_TEST_IMAGE"/></xdr:nvPicPr>
                  <xdr:blipFill><a:blip r:embed="rId1"/></xdr:blipFill>
                </xdr:pic>
              </etc:cellImage>
            </etc:cellImages>
            """,
        )
        archive.writestr(
            "xl/_rels/cellimages.xml.rels",
            """
            <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
              <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/proof.png"/>
            </Relationships>
            """,
        )
    manifest_dir = tmp_path / "03_format_probe"
    manifest_dir.mkdir()
    write_jsonl(
        manifest_dir / "probed_manifest.jsonl",
        [
            {
                "file_id": "file_abc",
                "source_path": str(source_workbook),
                "file_name": "source.xlsx",
            }
        ],
    )

    prediction = make_prediction(
        "field_1",
        "C6",
        "双路市电",
        status="partial_clue",
        reference_docs=[],
        evidence_ids=["att_file_abc_E3_dispimg"],
    )

    patch_workbook(
        template,
        [prediction],
        output,
        overlays_by_field_id={"field_1": {"writeback_allowed": False, "critic_flags": [], "review_required": True}},
        writeback_config={"allow_uncertain": True},
        run_id="run_1",
    )

    workbook = load_workbook(output)
    assert_adjacent_evidence_layout(workbook, image_count=1)
    assert "att_file_abc_E3_dispimg" in workbook["Sheet1"]["D6"].value
