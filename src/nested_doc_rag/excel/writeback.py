from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from nested_doc_rag.io import read_json, read_jsonl, write_json, write_jsonl
from nested_doc_rag.schemas.eval import FieldPrediction
from nested_doc_rag.schemas.excel import ExcelWritebackItem, ReviewItem, WritebackAuditRecord, WritebackSummary

Mode = Literal["safe", "overwrite"]

MAX_EXCEL_ROW = 1_048_576
MAX_EXCEL_COLUMN = 16_384
COMMENT_AUTHOR = "nested-doc-rag"
WRITEBACK_STATUSES = {"confirmed", "uncertain", "flagged"}
WRITEBACK_ACTIONS = {
    "written",
    "written_red_comment",
    "review_only",
    "skipped_uncertain_policy",
    "skipped_non_empty_cell",
    "skipped_formula",
    "invalid_cell",
    "duplicate_target_cell",
}
WB_INVALID_CELL = "WB_INVALID_CELL"
WB_MISSING_EVIDENCE = "WB_MISSING_EVIDENCE"
WB_POLICY_REJECTED = "WB_POLICY_REJECTED"
WB_COMMENT_TOO_LONG = "WB_COMMENT_TOO_LONG"
CRITICAL_FLAGS = {
    "answered_without_source",
    "invalid_source_reference",
    "answered_from_global_intro_risk",
    "answer_too_long",
    "scope_mismatch_risk",
    "liquid_cooling_scope_mismatch",
    "field_intent_source_mismatch",
}
SUPPORTED_EVIDENCE_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff"}
DEFAULT_EVIDENCE_IMAGE_MAX_WIDTH = 360
DEFAULT_EVIDENCE_IMAGE_MAX_HEIGHT = 240


@dataclass(frozen=True)
class TargetCell:
    sheet_name: str
    cell: str
    key: str


@dataclass(frozen=True)
class ResolvedPrediction:
    prediction: FieldPrediction
    target: TargetCell
    status: str
    evidence_refs: list[dict[str, Any]]


@dataclass(frozen=True)
class WritebackPolicy:
    allow_uncertain: bool = False
    uncertain_style: str = "red_fill"
    uncertain_comment_prefix: str = "[UNCERTAIN]"
    embed_evidence_images: bool = True
    evidence_image_mode: str = "adjacent_columns"
    max_evidence_images_per_field: int = 3
    max_comment_chars: int = 2000
    evidence_image_max_width_px: int = DEFAULT_EVIDENCE_IMAGE_MAX_WIDTH
    evidence_image_max_height_px: int = DEFAULT_EVIDENCE_IMAGE_MAX_HEIGHT

    @classmethod
    def from_value(cls, value: Any | None) -> WritebackPolicy:
        if value is None:
            return cls()
        if isinstance(value, Mapping):
            getter = value.get
        else:
            def getter(name: str, default: Any = None) -> Any:
                return getattr(value, name, default)

        return cls(
            allow_uncertain=bool(getter("allow_uncertain", False)),
            uncertain_style=str(getter("uncertain_style", "red_fill")),
            uncertain_comment_prefix=str(getter("uncertain_comment_prefix", "[UNCERTAIN]")),
            embed_evidence_images=bool(getter("embed_evidence_images", True)),
            evidence_image_mode=str(getter("evidence_image_mode", "adjacent_columns")),
            max_evidence_images_per_field=int(getter("max_evidence_images_per_field", 3) or 3),
            max_comment_chars=int(getter("max_comment_chars", 2000) or 2000),
            evidence_image_max_width_px=int(getter("evidence_image_max_width_px", DEFAULT_EVIDENCE_IMAGE_MAX_WIDTH) or DEFAULT_EVIDENCE_IMAGE_MAX_WIDTH),
            evidence_image_max_height_px=int(getter("evidence_image_max_height_px", DEFAULT_EVIDENCE_IMAGE_MAX_HEIGHT) or DEFAULT_EVIDENCE_IMAGE_MAX_HEIGHT),
        )


def prepare_writeback_item(sheet_name: str, cell: str, value: object, comment: str | None = None) -> ExcelWritebackItem:
    return ExcelWritebackItem(sheet_name=sheet_name, cell=cell, value=value, comment=comment)


def writeback_from_files(
    *,
    template_path: Path,
    predictions_path: Path,
    output_path: Path,
    trace_path: Path | None = None,
    evidence_map_path: Path | None = None,
    mode: Mode = "safe",
    write_comments: bool = True,
) -> WritebackSummary:
    predictions = [FieldPrediction.from_dict(record) for record in read_jsonl(predictions_path)]
    trace_by_field = load_trace_index(trace_path) if trace_path else {}
    evidence_map = read_json(evidence_map_path) if evidence_map_path and evidence_map_path.exists() else None
    return patch_workbook(
        template_path=template_path,
        predictions=predictions,
        output_path=output_path,
        mode=mode,
        write_comments=write_comments,
        trace_by_field=trace_by_field,
        evidence_map=evidence_map,
    )


def patch_workbook(
    template_path: Path,
    predictions: list[FieldPrediction],
    output_path: Path,
    mode: Mode = "safe",
    write_comments: bool = True,
    *,
    trace_by_field: dict[str, str] | None = None,
    evidence_map: dict[str, Any] | None = None,
    overlays_by_field_id: Mapping[str, Any] | None = None,
    writeback_config: Any | None = None,
    run_id: str | None = None,
) -> WritebackSummary:
    if mode not in {"safe", "overwrite"}:
        raise ValueError("mode must be 'safe' or 'overwrite'")

    workbook = load_workbook(template_path)
    trace_by_field = trace_by_field or {}
    overlays_by_field_id = overlays_by_field_id or {}
    policy = WritebackPolicy.from_value(writeback_config)
    audit_records: list[WritebackAuditRecord] = []
    review_items: list[ReviewItem] = []
    image_evidence_records: list[dict[str, Any]] = []
    evidence_output: dict[str, Any] = dict(evidence_map or {})
    evidence_fields = dict(evidence_output.get("fields") or {})
    evidence_output["fields"] = evidence_fields

    resolved: list[ResolvedPrediction] = []
    for prediction in predictions:
        overlay = overlays_by_field_id.get(prediction.field_id)
        status = classify_writeback_status(prediction, overlay)
        evidence_refs = evidence_refs_for_prediction(
            prediction,
            overlay,
            run_id=run_id,
            max_image_refs=policy.max_evidence_images_per_field,
        )
        image_evidence_records.extend(image_evidence_from_refs(prediction.field_id, evidence_refs))
        add_evidence_record(evidence_fields, prediction, trace_by_field.get(prediction.field_id), status=status, evidence_refs=evidence_refs)
        target = resolve_target_cell(workbook, prediction.target_cell)
        if isinstance(target, str):
            audit_records.append(
                make_audit(
                    prediction,
                    action="invalid",
                    reason=target,
                    trace_id=trace_by_field.get(prediction.field_id),
                    status="flagged",
                    writeback_action="invalid_cell",
                    evidence_refs=evidence_refs,
                    error_code=WB_INVALID_CELL,
                )
            )
            review_items.append(
                make_review_item(
                    prediction,
                    reason=target,
                    trace_id=trace_by_field.get(prediction.field_id),
                    status="flagged",
                    writeback_action="invalid_cell",
                    evidence_refs=evidence_refs,
                    error_code=WB_INVALID_CELL,
                )
            )
            continue
        resolved.append(ResolvedPrediction(prediction=prediction, target=target, status=status, evidence_refs=evidence_refs))

    duplicate_keys = {item.target.key for item in resolved if sum(1 for other in resolved if other.target.key == item.target.key) > 1}
    for item in resolved:
        prediction = item.prediction
        trace_id = trace_by_field.get(prediction.field_id)
        if item.target.key in duplicate_keys:
            audit_records.append(
                make_audit(
                    prediction,
                    action="conflict",
                    reason="duplicate_target_cell",
                    target=item.target,
                    trace_id=trace_id,
                    status="flagged",
                    writeback_action="duplicate_target_cell",
                    evidence_refs=item.evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                )
            )
            review_items.append(
                make_review_item(
                    prediction,
                    reason="duplicate_target_cell",
                    trace_id=trace_id,
                    status="flagged",
                    writeback_action="duplicate_target_cell",
                    evidence_refs=item.evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                )
            )
            continue

        worksheet = workbook[item.target.sheet_name]
        cell = worksheet[item.target.cell]
        audit_record, review_item = apply_prediction(
            cell=cell,
            prediction=prediction,
            target=item.target,
            mode=mode,
            write_comments=write_comments,
            trace_id=trace_id,
            status=item.status,
            evidence_refs=item.evidence_refs,
            policy=policy,
        )
        audit_records.append(audit_record)
        if review_item:
            review_items.append(review_item)

    write_evidence_outputs(workbook, audit_records, policy=policy, base_dir=output_path.parent)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    audit_path = output_path.parent / "writeback_audit.jsonl"
    evidence_map_output_path = output_path.parent / "evidence_map.json"
    review_items_path = output_path.parent / "review_items.jsonl"
    image_evidence_path = output_path.parent / "image_evidence.jsonl"
    write_jsonl(audit_path, [record.to_dict() for record in audit_records])
    write_json(evidence_map_output_path, evidence_output)
    write_jsonl(review_items_path, [item.to_dict() for item in review_items])
    if image_evidence_records:
        write_jsonl(image_evidence_path, image_evidence_records)
    return build_summary(
        output_path=output_path,
        audit_path=audit_path,
        evidence_map_path=evidence_map_output_path,
        review_items_path=review_items_path,
        image_evidence_path=image_evidence_path if image_evidence_records else None,
        audit_records=audit_records,
        review_items=review_items,
    )


def apply_prediction(
    *,
    cell: Cell | MergedCell,
    prediction: FieldPrediction,
    target: TargetCell,
    mode: Mode,
    write_comments: bool,
    trace_id: str | None,
    status: str,
    evidence_refs: list[dict[str, Any]],
    policy: WritebackPolicy,
) -> tuple[WritebackAuditRecord, ReviewItem | None]:
    if status == "flagged":
        reason = flagged_reason(prediction)
        return (
            make_audit(
                prediction,
                action="skipped",
                reason=reason,
                target=target,
                trace_id=trace_id,
                status=status,
                writeback_action="review_only",
                evidence_refs=evidence_refs,
            ),
            make_review_item(
                prediction,
                reason=reason,
                trace_id=trace_id,
                status=status,
                writeback_action="review_only",
                evidence_refs=evidence_refs,
            ),
        )

    if status == "uncertain" and not evidence_refs:
        return (
            make_audit(
                prediction,
                action="skipped",
                reason="missing_evidence",
                target=target,
                trace_id=trace_id,
                status=status,
                writeback_action="review_only",
                evidence_refs=evidence_refs,
                error_code=WB_MISSING_EVIDENCE,
            ),
            make_review_item(
                prediction,
                reason="missing_evidence",
                trace_id=trace_id,
                status=status,
                writeback_action="review_only",
                evidence_refs=evidence_refs,
                error_code=WB_MISSING_EVIDENCE,
            ),
        )

    if mode == "safe" and is_formula_cell(cell):
        return (
            make_audit(
                prediction,
                action="skipped",
                reason="skipped_formula",
                target=target,
                trace_id=trace_id,
                status="flagged" if status == "confirmed" else status,
                writeback_action="skipped_formula",
                evidence_refs=evidence_refs,
                error_code=WB_POLICY_REJECTED,
            ),
            make_review_item(
                prediction,
                reason="skipped_formula",
                trace_id=trace_id,
                status="flagged" if status == "confirmed" else status,
                writeback_action="skipped_formula",
                evidence_refs=evidence_refs,
                error_code=WB_POLICY_REJECTED,
            ),
        )

    if isinstance(cell, MergedCell):
        return (
            make_audit(
                prediction,
                action="invalid",
                reason="invalid_cell",
                target=target,
                trace_id=trace_id,
                status="flagged" if status == "confirmed" else status,
                writeback_action="invalid_cell",
                evidence_refs=evidence_refs,
                error_code=WB_INVALID_CELL,
            ),
            make_review_item(
                prediction,
                reason="invalid_cell",
                trace_id=trace_id,
                status="flagged" if status == "confirmed" else status,
                writeback_action="invalid_cell",
                evidence_refs=evidence_refs,
                error_code=WB_INVALID_CELL,
            ),
        )

    if status == "uncertain":
        if not policy.allow_uncertain:
            return (
                make_audit(
                    prediction,
                    action="skipped",
                    reason="uncertain_writeback_disabled",
                    target=target,
                    trace_id=trace_id,
                    status=status,
                    writeback_action="skipped_uncertain_policy",
                    evidence_refs=evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                ),
                make_review_item(
                    prediction,
                    reason="uncertain_writeback_disabled",
                    trace_id=trace_id,
                    status=status,
                    writeback_action="skipped_uncertain_policy",
                    evidence_refs=evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                ),
            )
        if not is_empty_cell(cell):
            return (
                make_audit(
                    prediction,
                    action="skipped",
                    reason="uncertain_target_not_empty",
                    target=target,
                    trace_id=trace_id,
                    status=status,
                    writeback_action="skipped_non_empty_cell",
                    evidence_refs=evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                ),
                make_review_item(
                    prediction,
                    reason="uncertain_target_not_empty",
                    trace_id=trace_id,
                    status=status,
                    writeback_action="skipped_non_empty_cell",
                    evidence_refs=evidence_refs,
                    error_code=WB_POLICY_REJECTED,
                ),
            )
        cell.value = prediction.answer_value
        apply_uncertain_style(cell, policy)
        comment_length, comment_truncated = maybe_write_uncertain_comment(
            cell,
            prediction,
            evidence_refs,
            write_comments=write_comments,
            trace_id=trace_id,
            policy=policy,
        )
        return (
            make_audit(
                prediction,
                action="written",
                reason="written_uncertain",
                target=target,
                trace_id=trace_id,
                status=status,
                writeback_action="written_red_comment",
                evidence_refs=evidence_refs,
                error_code=WB_COMMENT_TOO_LONG if comment_truncated else None,
                comment_length=comment_length,
            ),
            make_review_item(
                prediction,
                reason="uncertain_written",
                trace_id=trace_id,
                status=status,
                writeback_action="written_red_comment",
                evidence_refs=evidence_refs,
            ),
        )

    cell.value = prediction.answer_value
    comment_length, comment_truncated = maybe_write_comment(
        cell,
        prediction,
        evidence_refs,
        write_comments=write_comments,
        trace_id=trace_id,
        policy=policy,
    )
    return (
        make_audit(
            prediction,
            action="written",
            reason="written",
            target=target,
            trace_id=trace_id,
            status="confirmed",
            writeback_action="written",
            evidence_refs=evidence_refs,
            error_code=WB_COMMENT_TOO_LONG if comment_truncated else None,
            comment_length=comment_length,
        ),
        None,
    )


def maybe_write_comment(
    cell: Cell | MergedCell,
    prediction: FieldPrediction,
    evidence_refs: list[dict[str, Any]],
    *,
    write_comments: bool,
    trace_id: str | None,
    policy: WritebackPolicy,
) -> tuple[int, bool]:
    if not write_comments or isinstance(cell, MergedCell):
        return 0, False
    comment_text = build_evidence_comment(prediction, evidence_refs, trace_id=trace_id)
    return write_limited_comment(cell, comment_text, policy=policy)


def maybe_write_uncertain_comment(
    cell: Cell | MergedCell,
    prediction: FieldPrediction,
    evidence_refs: list[dict[str, Any]],
    *,
    write_comments: bool,
    trace_id: str | None,
    policy: WritebackPolicy,
) -> tuple[int, bool]:
    if not write_comments or isinstance(cell, MergedCell):
        return 0, False
    comment_text = build_evidence_comment(
        prediction,
        evidence_refs,
        trace_id=trace_id,
        prefix=policy.uncertain_comment_prefix,
    )
    return write_limited_comment(cell, comment_text, policy=policy)


def write_limited_comment(cell: Cell, comment_text: str, *, policy: WritebackPolicy) -> tuple[int, bool]:
    if cell.comment and cell.comment.text:
        comment_text = f"{cell.comment.text}\n\n{comment_text}"
    truncated = False
    limit = max(0, policy.max_comment_chars)
    if len(comment_text) > limit:
        suffix = f"\n{WB_COMMENT_TOO_LONG}: comment truncated"
        if limit == 0:
            comment_text = ""
        elif limit <= len(suffix):
            comment_text = suffix[-limit:]
        else:
            comment_text = comment_text[: limit - len(suffix)].rstrip() + suffix
        truncated = True
    cell.comment = Comment(comment_text, COMMENT_AUTHOR)
    return len(comment_text), truncated


def build_evidence_comment(
    prediction: FieldPrediction,
    evidence_refs: list[dict[str, Any]],
    *,
    trace_id: str | None,
    prefix: str | None = None,
) -> str:
    lines = []
    if prefix:
        lines.append(prefix)
    lines.append("evidence:")
    for index, ref in enumerate(evidence_refs[:5], start=1):
        lines.append(f"{index}. document: {evidence_document_name(ref)}")
        text = evidence_text(ref)
        if text:
            lines.append(f"   text: {display_comment_value(text, max_chars=500)}")
    if len(lines) == 1:
        lines.append(f"1. document: {format_list(prediction.source_chunk_ids) or 'unknown'}")
    return "\n".join(lines)


def evidence_document_name(ref: Mapping[str, Any]) -> str:
    return str(
        ref.get("file_name")
        or ref.get("document_name")
        or ref.get("document_id")
        or ref.get("object_key")
        or ref.get("chunk_id")
        or "unknown"
    )


def evidence_text(ref: Mapping[str, Any]) -> str:
    return str(
        ref.get("text_preview")
        or ref.get("raw_text")
        or ref.get("text")
        or ref.get("content")
        or ref.get("snippet")
        or ""
    ).strip()


def format_list(values: list[Any]) -> str:
    return ", ".join(str(value) for value in values)


def display_comment_value(value: Any, max_chars: int = 180) -> str:
    text = "" if value is None else str(value).strip()
    return text if len(text) <= max_chars else text[: max_chars - 3] + "..."


def ref_location(ref: Mapping[str, Any]) -> str:
    parts = [
        ref.get("source_anchor"),
        f"page {ref.get('page')}" if ref.get("page") is not None else None,
        ref.get("sheet_name"),
        ref.get("cell"),
    ]
    return " / ".join(str(part) for part in parts if part) or "unknown"


def apply_uncertain_style(cell: Cell, policy: WritebackPolicy) -> None:
    if policy.uncertain_style == "red_font":
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            bold=cell.font.bold,
            italic=cell.font.italic,
            color="FFFF0000",
        )
        return
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFCCCC")


def is_empty_cell(cell: Cell | MergedCell) -> bool:
    return cell.value is None or str(cell.value).strip() == ""


def validation_passed(prediction: FieldPrediction) -> bool:
    validation = prediction.validation or {}
    if validation.get("validation_pass") is False or validation.get("passed") is False:
        return False
    if validation.get("validation_failed") or validation.get("evidence_missing"):
        return False
    for key in ("constraint_violations", "errors", "error_messages"):
        if validation.get(key):
            return False
    return True


def flagged_reason(prediction: FieldPrediction) -> str:
    if prediction.answer_status != "answered":
        return "skipped_status"
    if not validation_passed(prediction):
        return "validation_failed"
    return "review_only"


def classify_writeback_status(prediction: FieldPrediction, overlay: Any | None) -> str:
    writeback_allowed = bool(overlay_value(overlay, "writeback_allowed", False))
    critic_flags = {str(flag) for flag in overlay_value(overlay, "critic_flags", []) or []}
    if overlay is None and prediction.answer_status == "answered" and validation_passed(prediction):
        return "confirmed"
    if prediction.answer_status == "answered" and validation_passed(prediction) and writeback_allowed:
        return "confirmed"
    if has_usable_answer(prediction) and has_evidence(prediction, overlay) and not critic_flags.intersection(CRITICAL_FLAGS):
        return "uncertain"
    return "flagged"


def has_usable_answer(prediction: FieldPrediction) -> bool:
    if prediction.answer_status not in {"answered", "partial_clue"}:
        return False
    text = display_comment_value(prediction.answer_value, max_chars=400)
    if not text:
        return False
    blocked_prefixes = ("未找到", "无可直接填写", "检索到相关线索，但证据不足")
    return not any(text.startswith(prefix) for prefix in blocked_prefixes)


def has_evidence(prediction: FieldPrediction, overlay: Any | None) -> bool:
    return bool(
        prediction.source_chunk_ids
        or prediction.reference_chunk_ids
        or prediction.reference_source_documents
        or prediction.evidence_attachment_ids
        or overlay_value(overlay, "suggested_reference_source_documents", [])
        or overlay_value(overlay, "suggested_reference_chunk_ids", [])
    )


def overlay_value(overlay: Any | None, name: str, default: Any = None) -> Any:
    if overlay is None:
        return default
    if isinstance(overlay, Mapping):
        return overlay.get(name, default)
    return getattr(overlay, name, default)


def is_formula_cell(cell: Cell | MergedCell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def resolve_target_cell(workbook: Workbook, target_cell: str | None) -> TargetCell | str:
    if not target_cell:
        return "invalid_cell"
    raw = str(target_cell).strip()
    if not raw:
        return "invalid_cell"

    if "!" in raw:
        sheet_name_raw, cell_ref = raw.rsplit("!", 1)
        sheet_name = unquote_sheet_name(sheet_name_raw)
    else:
        sheet_name = workbook.active.title
        cell_ref = raw

    if sheet_name not in workbook.sheetnames:
        return "invalid_cell"

    coordinate = normalize_coordinate(cell_ref)
    if not coordinate:
        return "invalid_cell"

    worksheet = workbook[sheet_name]
    coordinate = resolve_merged_coordinate(worksheet, coordinate)
    return TargetCell(sheet_name=sheet_name, cell=coordinate, key=f"{sheet_name}!{coordinate}")


def unquote_sheet_name(value: str) -> str:
    text = value.strip()
    if len(text) >= 2 and text[0] == "'" and text[-1] == "'":
        return text[1:-1].replace("''", "'")
    return text


def normalize_coordinate(value: str) -> str | None:
    try:
        column_letters, row = coordinate_from_string(value.strip())
        column = column_index_from_string(column_letters)
    except ValueError:
        return None
    if row < 1 or row > MAX_EXCEL_ROW or column < 1 or column > MAX_EXCEL_COLUMN:
        return None
    return f"{get_column_letter(column)}{row}"


def resolve_merged_coordinate(worksheet: Worksheet, coordinate: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
    return coordinate


def make_audit(
    prediction: FieldPrediction,
    *,
    action: str,
    reason: str,
    target: TargetCell | None = None,
    trace_id: str | None = None,
    status: str = "flagged",
    writeback_action: str = "review_only",
    evidence_refs: list[dict[str, Any]] | None = None,
    error_code: str | None = None,
    comment_length: int = 0,
) -> WritebackAuditRecord:
    evidence_refs = evidence_refs or []
    return WritebackAuditRecord(
        field_id=prediction.field_id,
        row_index=prediction.row_index,
        target_cell=prediction.target_cell,
        sheet_name=target.sheet_name if target else None,
        cell=target.cell if target else None,
        action=action,
        reason=reason,
        answer_status=prediction.answer_status,
        confidence=prediction.confidence,
        answer_value=prediction.answer_value,
        source_chunk_ids=prediction.source_chunk_ids,
        evidence_attachment_ids=prediction.evidence_attachment_ids,
        trace_id=trace_id,
        status=status,
        writeback_action=writeback_action,
        evidence_refs=evidence_refs,
        evidence_count=len(evidence_refs),
        image_evidence_count=sum(1 for ref in evidence_refs if ref.get("image_object_key")),
        error_code=error_code,
        comment_length=comment_length,
    )


def make_review_item(
    prediction: FieldPrediction,
    *,
    reason: str,
    trace_id: str | None = None,
    status: str = "flagged",
    writeback_action: str = "review_only",
    evidence_refs: list[dict[str, Any]] | None = None,
    error_code: str | None = None,
) -> ReviewItem:
    return ReviewItem(
        field_id=prediction.field_id,
        row_index=prediction.row_index,
        target_cell=prediction.target_cell,
        reason=reason,
        answer_status=prediction.answer_status,
        confidence=prediction.confidence,
        answer_value=prediction.answer_value,
        source_chunk_ids=prediction.source_chunk_ids,
        evidence_attachment_ids=prediction.evidence_attachment_ids,
        trace_id=trace_id,
        status=status,
        writeback_action=writeback_action,
        evidence_refs=evidence_refs or [],
        error_code=error_code,
    )


def add_evidence_record(
    evidence_fields: dict[str, Any],
    prediction: FieldPrediction,
    trace_id: str | None,
    *,
    status: str,
    evidence_refs: list[dict[str, Any]],
) -> None:
    evidence_fields[prediction.field_id] = {
        "target_cell": prediction.target_cell,
        "answer_status": prediction.answer_status,
        "confidence": prediction.confidence,
        "source_chunk_ids": prediction.source_chunk_ids,
        "evidence_attachment_ids": prediction.evidence_attachment_ids,
        "status": status,
        "evidence_refs": evidence_refs,
        "trace_id": trace_id,
    }


def evidence_refs_for_prediction(
    prediction: FieldPrediction,
    overlay: Any | None,
    *,
    run_id: str | None,
    max_image_refs: int,
) -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    docs.extend(dict(item) for item in prediction.reference_source_documents if isinstance(item, dict))
    docs.extend(dict(item) for item in overlay_value(overlay, "suggested_reference_source_documents", []) or [] if isinstance(item, dict))
    if not docs:
        docs.extend({"chunk_id": chunk_id} for chunk_id in prediction.source_chunk_ids)

    refs: list[dict[str, Any]] = []
    seen: set[str] = set()
    image_count = 0
    for doc in docs:
        ref = normalize_evidence_ref(doc)
        key = "|".join(str(ref.get(part) or "") for part in ("chunk_id", "source_anchor", "object_key", "qdrant_point_id"))
        if key in seen:
            continue
        seen.add(key)
        proof_attachments = normalize_proof_attachments(doc.get("proof_attachments"))
        proof_ids = [str(item) for item in doc.get("proof_attachment_ids") or [] if item]
        if not proof_ids:
            proof_ids = [str(item.get("attachment_id")) for item in proof_attachments if item.get("attachment_id")]
        if not ref.get("image_object_key") and proof_ids and image_count < max_image_refs:
            attach_proof_metadata(ref, proof_ids[0], proof_attachments)
            ref["image_object_key"] = ref.get("image_object_key") or image_object_key_for(run_id, prediction.field_id, proof_ids[0])
            ref["caption"] = ref.get("caption") or "proof attachment"
            image_count += 1
        refs.append(ref)
        for attachment_id in proof_ids[1:]:
            if image_count >= max_image_refs:
                break
            object_key = image_object_key_for(run_id, prediction.field_id, attachment_id)
            if any(ref_item.get("image_object_key") == object_key for ref_item in refs):
                continue
            proof_ref = dict(ref)
            attach_proof_metadata(proof_ref, attachment_id, proof_attachments)
            proof_ref["image_object_key"] = proof_ref.get("image_object_key") or object_key
            proof_ref["caption"] = proof_ref.get("caption") or "proof attachment"
            refs.append(proof_ref)
            image_count += 1

    if prediction.evidence_attachment_ids:
        for attachment_id in prediction.evidence_attachment_ids:
            if image_count >= max_image_refs:
                break
            object_key = image_object_key_for(run_id, prediction.field_id, attachment_id)
            if any(ref.get("image_object_key") == object_key for ref in refs):
                continue
            refs.append(
                {
                    "chunk_id": "",
                    "document_id": "",
                    "object_key": "",
                    "qdrant_point_id": "",
                    "source_type": "attachment",
                    "source_anchor": "",
                    "page": None,
                    "sheet_name": "",
                    "cell": "",
                    "image_object_key": object_key,
                    "proof_attachment_id": attachment_id,
                    "bbox": [],
                    "caption": "evidence attachment",
                }
            )
            image_count += 1
    return refs


def normalize_evidence_ref(doc: Mapping[str, Any]) -> dict[str, Any]:
    proof_attachments = normalize_proof_attachments(doc.get("proof_attachments"))
    return {
        "chunk_id": str(doc.get("chunk_id") or ""),
        "document_id": doc.get("document_id") or doc.get("doc_id") or doc.get("file_id") or "",
        "document_name": doc.get("document_name") or doc.get("source_document_name") or "",
        "object_key": doc.get("object_key") or doc.get("source_object_key") or "",
        "object_version_id": doc.get("object_version_id") or "",
        "qdrant_point_id": doc.get("qdrant_point_id") or doc.get("point_id") or "",
        "source_type": doc.get("source_type") or "",
        "source_anchor": doc.get("source_anchor") or doc.get("anchor") or "",
        "relative_path": doc.get("relative_path") or "",
        "source_file_path": doc.get("source_file_path") or doc.get("workbook_path") or doc.get("document_path") or "",
        "page": doc.get("page"),
        "sheet_name": doc.get("sheet_name") or "",
        "cell": doc.get("cell") or doc.get("cell_range") or "",
        "image_object_key": doc.get("image_object_key") or "",
        "image_path": first_text_value(
            doc,
            "image_path",
            "image_file_path",
            "local_image_path",
            "preview_image_path",
            "artifact_path",
            "embedded_payload_path",
        ),
        "media_path": doc.get("media_path") or "",
        "media_content_type": doc.get("media_content_type") or "",
        "proof_attachment_id": doc.get("proof_attachment_id") or doc.get("attachment_id") or "",
        "proof_attachments": proof_attachments,
        "bbox": doc.get("bbox") or [],
        "caption": doc.get("caption") or "",
        "file_name": doc.get("file_name") or doc.get("filename") or doc.get("source_file_name") or "",
        "text_preview": doc.get("text_preview") or doc.get("raw_text") or doc.get("text") or doc.get("content") or doc.get("snippet") or "",
    }


def normalize_proof_attachments(value: Any) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for item in value or []:
        if not isinstance(item, Mapping):
            continue
        output.append(
            {
                "attachment_id": item.get("attachment_id") or item.get("id") or "",
                "source_cell": item.get("source_cell") or item.get("cell") or "",
                "media_path": item.get("media_path") or "",
                "media_content_type": item.get("media_content_type") or item.get("content_type") or "",
                "attachment_type": item.get("attachment_type") or "",
                "image_id": item.get("image_id") or "",
                "relationship_id": item.get("relationship_id") or "",
                "image_path": first_text_value(
                    item,
                    "image_path",
                    "image_file_path",
                    "local_image_path",
                    "preview_image_path",
                    "artifact_path",
                    "embedded_payload_path",
                ),
                "caption": item.get("caption") or item.get("evidence_role") or "",
            }
        )
    return output


def attach_proof_metadata(ref: dict[str, Any], attachment_id: str, proof_attachments: list[dict[str, Any]]) -> None:
    ref["proof_attachment_id"] = attachment_id
    if proof_attachments and not ref.get("proof_attachments"):
        ref["proof_attachments"] = proof_attachments
    for attachment in proof_attachments:
        if attachment.get("attachment_id") != attachment_id:
            continue
        for key in ("media_path", "media_content_type", "image_path"):
            if attachment.get(key) and not ref.get(key):
                ref[key] = attachment[key]
        if attachment.get("source_cell") and not ref.get("cell"):
            ref["cell"] = attachment["source_cell"]
        if attachment.get("caption") and not ref.get("caption"):
            ref["caption"] = attachment["caption"]
        return
    parsed_cell = cell_from_attachment_id(attachment_id)
    if parsed_cell and not ref.get("cell"):
        ref["cell"] = parsed_cell


def first_text_value(value: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        raw = value.get(key)
        if raw:
            return str(raw)
    return ""


def image_object_key_for(run_id: str | None, field_id: str, attachment_id: str) -> str:
    safe_run = safe_object_key_part(run_id or "local_run")
    safe_field = safe_object_key_part(field_id)
    safe_attachment = safe_object_key_part(attachment_id)
    return f"runs/{safe_run}/evidence/{safe_field}/{safe_attachment}"


def safe_object_key_part(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_.=-]+", "_", str(value or "").strip())
    return text.strip("._") or "unknown"


def image_evidence_from_refs(field_id: str, evidence_refs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for ref in evidence_refs:
        image_object_key = ref.get("image_object_key")
        if not image_object_key:
            continue
        records.append(
            {
                "field_id": field_id,
                "image_object_key": image_object_key,
                "document_id": ref.get("document_id") or "",
                "source_anchor": ref.get("source_anchor") or "",
                "caption": ref.get("caption") or "",
                "proof_attachment_id": ref.get("proof_attachment_id") or "",
                "image_path": ref.get("image_path") or "",
                "media_path": ref.get("media_path") or "",
            }
        )
    return records


def write_evidence_outputs(
    workbook: Workbook,
    audit_records: list[WritebackAuditRecord],
    *,
    policy: WritebackPolicy | None = None,
    base_dir: Path | None = None,
) -> None:
    policy = policy or WritebackPolicy()
    if policy.evidence_image_mode == "append_sheet":
        write_evidence_sheet(workbook, audit_records, policy=policy, base_dir=base_dir)
        return
    if policy.evidence_image_mode in {"adjacent_columns", "inline", "answer_adjacent"}:
        remove_generated_evidence_sheet(workbook)
        write_adjacent_evidence_columns(workbook, audit_records, policy=policy, base_dir=base_dir)
        return
    remove_generated_evidence_sheet(workbook)
    write_adjacent_evidence_columns(workbook, audit_records, policy=policy, base_dir=base_dir)


def write_adjacent_evidence_columns(
    workbook: Workbook,
    audit_records: list[WritebackAuditRecord],
    *,
    policy: WritebackPolicy,
    base_dir: Path | None,
) -> None:
    records_by_sheet: dict[str, list[WritebackAuditRecord]] = {}
    for record in audit_records:
        if not record.sheet_name or not record.cell or not record.evidence_refs:
            continue
        if record.sheet_name not in workbook.sheetnames:
            continue
        records_by_sheet.setdefault(record.sheet_name, []).append(record)

    for sheet_name, records in records_by_sheet.items():
        sheet = workbook[sheet_name]
        min_row_by_answer_col = min_target_row_by_answer_column(records)
        for answer_column, min_row in min_row_by_answer_col.items():
            write_adjacent_headers(sheet, answer_column=answer_column, header_row=max(1, min_row - 1))
        for record in records:
            write_adjacent_evidence_record(sheet, record, policy=policy, base_dir=base_dir)


def min_target_row_by_answer_column(records: list[WritebackAuditRecord]) -> dict[int, int]:
    output: dict[int, int] = {}
    for record in records:
        if not record.cell:
            continue
        try:
            column_letters, row = coordinate_from_string(record.cell)
            answer_column = column_index_from_string(column_letters)
        except ValueError:
            continue
        output[answer_column] = min(output.get(answer_column, row), row)
    return output


def write_adjacent_headers(sheet: Worksheet, *, answer_column: int, header_row: int) -> None:
    evidence_column = answer_column + 1
    image_column = answer_column + 2
    if image_column > MAX_EXCEL_COLUMN:
        return
    evidence_header = sheet.cell(row=header_row, column=evidence_column)
    image_header = sheet.cell(row=header_row, column=image_column)
    if is_empty_cell(evidence_header):
        evidence_header.value = "证据"
    if is_empty_cell(image_header):
        image_header.value = "证据图片"
    evidence_header.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    image_header.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    sheet.column_dimensions[get_column_letter(evidence_column)].width = max(
        sheet.column_dimensions[get_column_letter(evidence_column)].width or 0,
        72,
    )
    sheet.column_dimensions[get_column_letter(image_column)].width = max(
        sheet.column_dimensions[get_column_letter(image_column)].width or 0,
        min(70, max(32, DEFAULT_EVIDENCE_IMAGE_MAX_WIDTH / 7)),
    )


def write_adjacent_evidence_record(
    sheet: Worksheet,
    record: WritebackAuditRecord,
    *,
    policy: WritebackPolicy,
    base_dir: Path | None,
) -> None:
    if not record.cell:
        return
    try:
        column_letters, row = coordinate_from_string(record.cell)
        answer_column = column_index_from_string(column_letters)
    except ValueError:
        return
    evidence_column = answer_column + 1
    image_column = answer_column + 2
    if image_column > MAX_EXCEL_COLUMN:
        return

    evidence_cell = sheet.cell(row=row, column=evidence_column)
    evidence_cell.value = build_adjacent_evidence_text(record)
    evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")

    line_count = max(1, str(evidence_cell.value or "").count("\n") + 1)
    text_height = min(180, max(30, line_count * 14))
    sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 0, text_height)

    proofs = collect_image_proofs([record], policy=policy)
    if not proofs:
        return
    if not policy.embed_evidence_images:
        image_cell = sheet.cell(row=row, column=image_column)
        image_cell.value = "\n".join(proof["image_object_key"] for proof in proofs if proof.get("image_object_key"))
        image_cell.alignment = Alignment(wrap_text=True, vertical="top")
        return

    inserted_count = 0
    unavailable: list[str] = []
    offset_y_px = 0
    for proof in proofs:
        inserted, image_width, image_height = insert_evidence_image(
            sheet,
            row=row,
            column=image_column,
            offset_y_px=offset_y_px,
            proof=proof,
            policy=policy,
            base_dir=base_dir,
        )
        if inserted:
            inserted_count += 1
            offset_y_px += image_height + 8
        else:
            unavailable.append(proof.get("image_object_key") or proof.get("proof_attachment_id") or proof.get("source") or "unknown")

    image_cell = sheet.cell(row=row, column=image_column)
    if unavailable:
        prefix = "" if inserted_count == 0 else f"已嵌入 {inserted_count} 张\n"
        image_cell.value = prefix + "图片不可用: " + ", ".join(unavailable)
        image_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if inserted_count:
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 0, offset_y_px * 0.75)


def build_adjacent_evidence_text(record: WritebackAuditRecord) -> str:
    lines = [
        f"状态: {record.status}",
        f"写回: {record.writeback_action}",
    ]
    if record.reason:
        lines.append(f"原因: {record.reason}")
    for index, ref in enumerate(record.evidence_refs, start=1):
        lines.append(f"{index}. 来源: {evidence_document_name(ref)}")
        location = ref_location(ref)
        if location != "unknown":
            lines.append(f"   位置: {location}")
        text = evidence_text(ref) or str(ref.get("caption") or "").strip()
        if text:
            lines.append(f"   证据: {display_comment_value(text, max_chars=800)}")
        if ref.get("image_object_key") or ref.get("proof_attachment_id"):
            lines.append(f"   图片: {ref.get('image_object_key') or ref.get('proof_attachment_id')}")
    return "\n".join(lines)


def remove_generated_evidence_sheet(workbook: Workbook) -> None:
    if "Evidence" in workbook.sheetnames:
        del workbook["Evidence"]


def write_evidence_sheet(
    workbook: Workbook,
    audit_records: list[WritebackAuditRecord],
    *,
    policy: WritebackPolicy | None = None,
    base_dir: Path | None = None,
) -> None:
    policy = policy or WritebackPolicy()
    rows: list[list[Any]] = []
    for record in audit_records:
        for ref in record.evidence_refs:
            rows.append(
                [
                    record.field_id,
                    record.status,
                    record.answer_value,
                    ref.get("file_name") or ref.get("document_id") or ref.get("object_key"),
                    ref_location(ref),
                    ref.get("text_preview") or ref.get("caption"),
                    ref.get("image_object_key") or "",
                    ref.get("image_path") or "",
                    ref.get("media_path") or "",
                ]
            )
    if not rows:
        return
    if "Evidence" in workbook.sheetnames:
        del workbook["Evidence"]
    sheet = workbook.create_sheet("Evidence")
    headers = ["field_id", "status", "answer_value", "source", "location", "text_preview", "image_object_key", "image_path", "media_path"]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    append_evidence_images(sheet, audit_records, policy=policy, base_dir=base_dir)


def append_evidence_images(
    sheet: Worksheet,
    audit_records: list[WritebackAuditRecord],
    *,
    policy: WritebackPolicy,
    base_dir: Path | None,
) -> None:
    proofs = collect_image_proofs(audit_records, policy=policy)
    if not proofs:
        return

    start_row = sheet.max_row + 2
    sheet.cell(row=start_row, column=1, value="Image Proofs")
    sheet.cell(row=start_row + 1, column=1, value="field_id")
    sheet.cell(row=start_row + 1, column=2, value="source")
    sheet.cell(row=start_row + 1, column=3, value="location")
    sheet.cell(row=start_row + 1, column=4, value="image_object_key")
    sheet.cell(row=start_row + 1, column=5, value="status")

    row = start_row + 2
    for proof in proofs:
        sheet.cell(row=row, column=1, value=proof["field_id"])
        sheet.cell(row=row, column=2, value=proof["source"])
        sheet.cell(row=row, column=3, value=proof["location"])
        sheet.cell(row=row, column=4, value=proof["image_object_key"])

        inserted = False
        if policy.embed_evidence_images:
            inserted = try_insert_evidence_image(sheet, row=row + 1, proof=proof, policy=policy, base_dir=base_dir)
        sheet.cell(row=row, column=5, value="inserted" if inserted else "image file unavailable")
        row += image_row_span(inserted, policy)


def collect_image_proofs(audit_records: list[WritebackAuditRecord], *, policy: WritebackPolicy) -> list[dict[str, Any]]:
    proofs: list[dict[str, Any]] = []
    seen: set[str] = set()
    for record in audit_records:
        per_field = 0
        for ref in record.evidence_refs:
            if not has_image_evidence(ref):
                continue
            key = str(ref.get("image_object_key") or ref.get("image_path") or ref.get("media_path") or f"{record.field_id}:{len(proofs)}")
            dedupe_key = f"{record.field_id}:{key}"
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            proofs.append(
                {
                    "field_id": record.field_id,
                    "source": evidence_document_name(ref),
                    "location": ref_location(ref),
                    "image_object_key": ref.get("image_object_key") or "",
                    "proof_attachment_id": ref.get("proof_attachment_id") or "",
                    "image_path": ref.get("image_path") or "",
                    "media_path": ref.get("media_path") or "",
                    "source_file_path": ref.get("source_file_path") or "",
                    "relative_path": ref.get("relative_path") or "",
                    "file_name": ref.get("file_name") or "",
                    "sheet_name": ref.get("sheet_name") or "",
                    "cell": ref.get("cell") or "",
                }
            )
            per_field += 1
            if per_field >= max(1, policy.max_evidence_images_per_field):
                break
    return proofs


def has_image_evidence(ref: Mapping[str, Any]) -> bool:
    return bool(ref.get("image_object_key") or ref.get("image_path") or ref.get("media_path"))


def try_insert_evidence_image(
    sheet: Worksheet,
    *,
    row: int,
    proof: Mapping[str, Any],
    policy: WritebackPolicy,
    base_dir: Path | None,
) -> bool:
    inserted, _, _ = insert_evidence_image(
        sheet,
        row=row,
        column=1,
        offset_y_px=0,
        proof=proof,
        policy=policy,
        base_dir=base_dir,
    )
    return inserted


def insert_evidence_image(
    sheet: Worksheet,
    *,
    row: int,
    column: int,
    offset_y_px: int,
    proof: Mapping[str, Any],
    policy: WritebackPolicy,
    base_dir: Path | None,
) -> tuple[bool, int, int]:
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except ImportError:
        return False, 0, 0

    image_source = open_image_source(proof, base_dir=base_dir)
    if image_source is None:
        return False, 0, 0

    try:
        image = OpenpyxlImage(image_source)
    except Exception:
        return False, 0, 0

    scale_image(image, max_width=policy.evidence_image_max_width_px, max_height=policy.evidence_image_max_height_px)
    if offset_y_px:
        image.anchor = image_cell_anchor(
            row=row,
            column=column,
            offset_y_px=offset_y_px,
            width_px=image.width,
            height_px=image.height,
        )
        sheet.add_image(image)
    else:
        sheet.add_image(image, f"{get_column_letter(column)}{row}")
    sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 0, image.height * 0.75)
    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width or 0, min(70, image.width / 7))
    return True, int(image.width), int(image.height)


def image_cell_anchor(*, row: int, column: int, offset_y_px: int, width_px: int, height_px: int) -> Any:
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    marker = AnchorMarker(col=column - 1, colOff=0, row=row - 1, rowOff=pixels_to_EMU(offset_y_px))
    ext = XDRPositiveSize2D(cx=pixels_to_EMU(width_px), cy=pixels_to_EMU(height_px))
    return OneCellAnchor(_from=marker, ext=ext)


def open_image_source(proof: Mapping[str, Any], *, base_dir: Path | None) -> str | BytesIO | None:
    image_path = resolve_existing_image_path(str(proof.get("image_path") or ""), base_dir=base_dir)
    if image_path:
        return str(image_path)

    registry_source = open_registry_image_source(proof, base_dir=base_dir)
    if registry_source is not None:
        return registry_source

    source_workbook = resolve_source_workbook(proof, base_dir=base_dir)
    media_path = str(proof.get("media_path") or "").strip()
    if source_workbook and media_path:
        return read_xlsx_media(source_workbook, media_path)
    if source_workbook:
        return read_xlsx_dispimg_media(
            source_workbook,
            sheet_name=str(proof.get("sheet_name") or ""),
            cell=dispimg_cell_reference(
                str(proof.get("cell") or ""),
                str(proof.get("proof_attachment_id") or ""),
            ),
        )
    return None


def open_registry_image_source(proof: Mapping[str, Any], *, base_dir: Path | None) -> str | BytesIO | None:
    attachment_id = str(proof.get("proof_attachment_id") or "")
    if not attachment_id:
        return None
    record = find_registry_record(attachment_id, base_dir=base_dir)
    if not record:
        return None
    image_path = resolve_existing_image_path(str(record.get("image_path") or ""), base_dir=base_dir)
    if image_path:
        return str(image_path)
    source_workbook = resolve_source_workbook(record, base_dir=base_dir)
    media_path = str(record.get("media_path") or "")
    if source_workbook and media_path:
        return read_xlsx_media(source_workbook, media_path)
    if source_workbook:
        return read_xlsx_dispimg_media(
            source_workbook,
            sheet_name=str(record.get("sheet_name") or ""),
            cell=dispimg_cell_reference(str(record.get("source_cell") or record.get("cell") or ""), attachment_id),
        )
    return None


def find_registry_record(attachment_id: str, *, base_dir: Path | None) -> dict[str, Any] | None:
    for registry_path in registry_candidate_paths(base_dir):
        try:
            rows = read_jsonl(registry_path)
        except Exception:
            continue
        for row in rows:
            if str(row.get("attachment_id") or "") == attachment_id:
                return dict(row)
    return None


def registry_candidate_paths(base_dir: Path | None) -> list[Path]:
    candidates: list[Path] = []
    if base_dir:
        candidates.extend(
            [
                base_dir / "proof_attachment_registry.jsonl",
                base_dir / "evidence_images" / "proof_attachment_registry.jsonl",
                base_dir.parent / "proof_attachment_registry.jsonl",
            ]
        )
    cwd = Path.cwd()
    candidates.extend(
        [
            cwd / "proof_attachment_registry.jsonl",
            cwd / "artifacts" / "11_embedding_build" / "proof_attachment_registry.jsonl",
        ]
    )
    artifacts = cwd / "artifacts"
    if artifacts.exists():
        candidates.extend(artifacts.glob("**/proof_attachment_registry.jsonl"))
    output: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            continue
        if resolved in seen or not resolved.exists() or not resolved.is_file():
            continue
        seen.add(resolved)
        output.append(resolved)
    return output


def resolve_existing_image_path(value: str, *, base_dir: Path | None) -> Path | None:
    if not value:
        return None
    raw = Path(value)
    candidates = [raw]
    if base_dir and not raw.is_absolute():
        candidates.append(base_dir / raw)
    if not raw.is_absolute():
        cwd = Path.cwd()
        candidates.extend([cwd / raw, cwd / "data" / raw, cwd / "artifacts" / raw])
    for candidate in candidates:
        try:
            resolved = candidate.expanduser().resolve()
        except OSError:
            continue
        if resolved.exists() and resolved.is_file() and is_supported_image_path(resolved):
            return resolved
    return None


def resolve_source_workbook(proof: Mapping[str, Any], *, base_dir: Path | None) -> Path | None:
    raw_values = [
        str(proof.get("source_file_path") or ""),
        str(proof.get("relative_path") or ""),
        str(proof.get("file_name") or ""),
    ]
    resolved = first_existing_workbook(raw_values, base_dir=base_dir)
    if resolved:
        return resolved
    return resolve_source_workbook_from_attachment(str(proof.get("proof_attachment_id") or ""), base_dir=base_dir)


def first_existing_workbook(raw_values: list[str], *, base_dir: Path | None) -> Path | None:
    candidates: list[Path] = []
    for raw_value in raw_values:
        if not raw_value:
            continue
        raw = Path(raw_value)
        candidates.append(raw)
        if base_dir and not raw.is_absolute():
            candidates.append(base_dir / raw)
        if not raw.is_absolute():
            cwd = Path.cwd()
            candidates.extend([cwd / raw, cwd / "data" / raw, cwd / "artifacts" / raw])
    for candidate in candidates:
        try:
            resolved = candidate.expanduser().resolve()
        except OSError:
            continue
        if resolved.exists() and resolved.is_file() and resolved.suffix.lower() in {".xlsx", ".xlsm"}:
            return resolved
    return None


def resolve_source_workbook_from_attachment(attachment_id: str, *, base_dir: Path | None) -> Path | None:
    file_id = file_id_from_attachment_id(attachment_id)
    if not file_id:
        return None
    for manifest_path in probed_manifest_candidate_paths(base_dir):
        try:
            rows = read_jsonl(manifest_path)
        except Exception:
            continue
        for row in rows:
            if str(row.get("file_id") or "") != file_id:
                continue
            return first_existing_workbook(
                [
                    str(row.get("source_path") or ""),
                    str(row.get("relative_path") or ""),
                    str(row.get("file_name") or ""),
                ],
                base_dir=base_dir,
            )
    return None


def file_id_from_attachment_id(attachment_id: str) -> str:
    match = re.search(r"att_(file_[0-9A-Fa-f]+)_", str(attachment_id or ""))
    return match.group(1) if match else ""


def probed_manifest_candidate_paths(base_dir: Path | None) -> list[Path]:
    candidates: list[Path] = []
    if base_dir:
        candidates.extend(
            [
                base_dir / "probed_manifest.jsonl",
                base_dir / "03_format_probe" / "probed_manifest.jsonl",
                base_dir.parent / "03_format_probe" / "probed_manifest.jsonl",
            ]
        )
    cwd = Path.cwd()
    candidates.append(cwd / "artifacts" / "03_format_probe" / "probed_manifest.jsonl")
    artifacts = cwd / "artifacts"
    if artifacts.exists():
        candidates.extend(artifacts.glob("**/probed_manifest.jsonl"))
    output: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            continue
        if resolved in seen or not resolved.exists() or not resolved.is_file():
            continue
        seen.add(resolved)
        output.append(resolved)
    return output


def read_xlsx_media(workbook_path: Path, media_path: str) -> BytesIO | None:
    normalized = media_path.replace("\\", "/").lstrip("/")
    if not normalized.startswith("xl/media/"):
        return None
    suffix = Path(normalized).suffix.lower()
    if suffix not in SUPPORTED_EVIDENCE_IMAGE_SUFFIXES:
        return None
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            data = archive.read(normalized)
    except (OSError, KeyError, zipfile.BadZipFile):
        return None
    return BytesIO(data)


def read_xlsx_dispimg_media(workbook_path: Path, *, sheet_name: str, cell: str) -> BytesIO | None:
    image_id = dispimg_image_id_from_cell(workbook_path, sheet_name=sheet_name, cell=cell)
    if not image_id:
        return None
    media_path = cell_image_media_path(workbook_path, image_id)
    if not media_path:
        return None
    return read_xlsx_media(workbook_path, media_path)


def dispimg_image_id_from_cell(workbook_path: Path, *, sheet_name: str, cell: str) -> str:
    coordinate = normalize_cell_reference(cell)
    if not coordinate:
        return ""
    try:
        workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    except Exception:
        return ""
    try:
        worksheets = [workbook[sheet_name]] if sheet_name and sheet_name in workbook.sheetnames else list(workbook.worksheets)
        for worksheet in worksheets:
            value = worksheet[coordinate].value
            if not isinstance(value, str):
                continue
            match = re.search(r'DISPIMG\("([^"]+)"', value)
            if match:
                return match.group(1)
    finally:
        workbook.close()
    return ""


def cell_image_media_path(workbook_path: Path, image_id: str) -> str:
    if not image_id:
        return ""
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            image_xml = archive.read("xl/cellimages.xml")
            rels_xml = archive.read("xl/_rels/cellimages.xml.rels")
    except (OSError, KeyError, zipfile.BadZipFile):
        return ""

    rels = parse_relationships(rels_xml)
    try:
        root = ET.fromstring(image_xml)
    except ET.ParseError:
        return ""
    namespaces = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    embed_key = f"{{{namespaces['r']}}}embed"
    for pic in root.findall(".//xdr:pic", namespaces):
        name_node = pic.find(".//xdr:cNvPr", namespaces)
        blip_node = pic.find(".//a:blip", namespaces)
        if name_node is None or blip_node is None:
            continue
        if name_node.attrib.get("name") != image_id:
            continue
        target = rels.get(blip_node.attrib.get(embed_key, ""))
        return normalize_xlsx_target(target)
    return ""


def parse_relationships(value: bytes) -> dict[str, str]:
    try:
        root = ET.fromstring(value)
    except ET.ParseError:
        return {}
    output: dict[str, str] = {}
    for relationship in root:
        rel_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target")
        if rel_id and target:
            output[rel_id] = target
    return output


def normalize_xlsx_target(target: str | None) -> str:
    if not target:
        return ""
    normalized = target.replace("\\", "/").lstrip("/")
    if normalized.startswith("../"):
        normalized = normalized[3:]
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def is_supported_image_path(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EVIDENCE_IMAGE_SUFFIXES


def scale_image(image: Any, *, max_width: int, max_height: int) -> None:
    if image.width <= 0 or image.height <= 0:
        return
    scale = min(max_width / image.width, max_height / image.height, 1.0)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)


def image_row_span(inserted: bool, policy: WritebackPolicy) -> int:
    if not inserted:
        return 1
    return max(8, int(policy.evidence_image_max_height_px / 18) + 2)


def cell_from_attachment_id(attachment_id: str) -> str:
    match = re.search(r"_([A-Z]{1,3}[0-9]{1,7})_dispimg$", str(attachment_id or ""))
    return match.group(1) if match else ""


def dispimg_cell_reference(cell: str, attachment_id: str) -> str:
    text = str(cell or "").strip()
    attachment_cell = cell_from_attachment_id(attachment_id)
    if text and ":" not in text and normalize_coordinate(text):
        return text
    if attachment_cell:
        return attachment_cell
    return text


def normalize_cell_reference(value: str) -> str | None:
    text = str(value or "").strip()
    if ":" in text:
        text = text.split(":", 1)[0]
    return normalize_coordinate(text)


def load_trace_index(trace_path: Path) -> dict[str, str]:
    output: dict[str, str] = {}
    for record in read_jsonl(trace_path):
        field_id = record.get("field_id") or (record.get("prediction") or {}).get("field_id")
        trace_id = record.get("trace_id") or record.get("id") or record.get("run_id")
        if field_id and trace_id:
            output[str(field_id)] = str(trace_id)
    return output


def build_summary(
    *,
    output_path: Path,
    audit_path: Path,
    evidence_map_path: Path,
    review_items_path: Path,
    image_evidence_path: Path | None,
    audit_records: list[WritebackAuditRecord],
    review_items: list[ReviewItem],
) -> WritebackSummary:
    return WritebackSummary(
        output_path=output_path,
        audit_path=audit_path,
        evidence_map_path=evidence_map_path,
        review_items_path=review_items_path,
        total_count=len(audit_records),
        written_count=sum(1 for record in audit_records if record.action == "written"),
        skipped_count=sum(1 for record in audit_records if record.action == "skipped"),
        conflict_count=sum(1 for record in audit_records if record.action == "conflict"),
        invalid_count=sum(1 for record in audit_records if record.reason == "invalid_cell"),
        formula_skipped_count=sum(1 for record in audit_records if record.reason == "skipped_formula"),
        review_count=len(review_items),
        confirmed_count=sum(1 for record in audit_records if record.status == "confirmed"),
        uncertain_count=sum(1 for record in audit_records if record.status == "uncertain"),
        flagged_count=sum(1 for record in audit_records if record.status == "flagged"),
        image_evidence_path=image_evidence_path,
        fields=[manifest_field_from_audit(record) for record in audit_records],
    )


def manifest_field_from_audit(record: WritebackAuditRecord) -> dict[str, Any]:
    return {
        "field_key": record.field_id,
        "field_id": record.field_id,
        "row_index": record.row_index,
        "target_cell": record.target_cell,
        "sheet_name": record.sheet_name,
        "cell": record.cell,
        "status": record.status,
        "answer_status": record.answer_status,
        "answer_value": record.answer_value,
        "writeback_action": record.writeback_action,
        "evidence_refs": record.evidence_refs,
        "error_code": record.error_code,
    }
