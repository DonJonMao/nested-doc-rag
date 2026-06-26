from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from collections.abc import Mapping
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

SUPPORTED_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff"}


def safe_key_part(value: Any) -> str:
    text = re.sub(r"[^A-Za-z0-9_.=-]+", "_", str(value or "").strip())
    return text.strip("._") or "unknown"


def file_id_for_path(path: Path | str) -> str:
    return "file_" + safe_key_part(str(path)).lower()[:48]


def attachment_id_for_dispimg(*, file_id: str, cell: str) -> str:
    return f"att_{safe_key_part(file_id)}_{safe_key_part(cell)}_dispimg"


def media_content_type(media_path: str) -> str:
    suffix = Path(media_path).suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".png":
        return "image/png"
    if suffix == ".gif":
        return "image/gif"
    if suffix == ".bmp":
        return "image/bmp"
    if suffix in {".tif", ".tiff"}:
        return "image/tiff"
    return "application/octet-stream"


def normalize_xlsx_target(target: str | None) -> str:
    if not target:
        return ""
    normalized = target.replace("\\", "/").lstrip("/")
    while normalized.startswith("../"):
        normalized = normalized[3:]
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def parse_relationships(value: bytes) -> dict[str, str]:
    try:
        root = ET.fromstring(value)
    except ET.ParseError:
        return {}
    output: dict[str, str] = {}
    for relationship in root:
        rel_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target")
        if rel_id and target:
            output[rel_id] = target
    return output


def cell_image_media_map(workbook_path: Path) -> dict[str, str]:
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            image_xml = archive.read("xl/cellimages.xml")
            rels_xml = archive.read("xl/_rels/cellimages.xml.rels")
    except (OSError, KeyError, zipfile.BadZipFile):
        return {}

    rels = parse_relationships(rels_xml)
    try:
        root = ET.fromstring(image_xml)
    except ET.ParseError:
        return {}

    namespaces = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    embed_key = f"{{{namespaces['r']}}}embed"
    output: dict[str, str] = {}
    for pic in root.findall(".//xdr:pic", namespaces):
        name_node = pic.find(".//xdr:cNvPr", namespaces)
        blip_node = pic.find(".//a:blip", namespaces)
        if name_node is None or blip_node is None:
            continue
        image_id = name_node.attrib.get("name")
        rel_id = blip_node.attrib.get(embed_key)
        media_path = normalize_xlsx_target(rels.get(rel_id or ""))
        if image_id and media_path:
            output[image_id] = media_path
    return output


def extract_dispimg_id(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    match = re.search(r'DISPIMG\("([^"]+)"', value)
    return match.group(1) if match else ""


def dispimg_cells(workbook_path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    except Exception:
        return []
    output: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    image_id = extract_dispimg_id(cell.value)
                    if not image_id:
                        continue
                    coordinate = f"{get_column_letter(cell.column)}{cell.row}"
                    output.append(
                        {
                            "sheet_name": worksheet.title,
                            "row_index": cell.row,
                            "cell": coordinate,
                            "image_id": image_id,
                        }
                    )
    finally:
        workbook.close()
    return output


def materialize_xlsx_dispimg_registry(
    workbook_path: Path,
    *,
    root: Path,
    output_dir: Path,
    namespace: str,
    knowledge_base_id: str,
) -> list[dict[str, Any]]:
    relative = relative_path(workbook_path, root)
    file_id = file_id_for_path(f"{knowledge_base_id}:{relative}")
    media_by_image_id = cell_image_media_map(workbook_path)
    rows: list[dict[str, Any]] = []
    for item in dispimg_cells(workbook_path):
        media_path = media_by_image_id.get(str(item["image_id"]) or "")
        attachment_id = attachment_id_for_dispimg(file_id=file_id, cell=str(item["cell"]))
        row = {
            "attachment_id": attachment_id,
            "file_id": file_id,
            "knowledge_base_id": knowledge_base_id,
            "namespace": namespace,
            "file_name": workbook_path.name,
            "relative_path": relative,
            "source_file_path": str(workbook_path.resolve()),
            "sheet_name": item["sheet_name"],
            "row_index": item["row_index"],
            "source_cell": item["cell"],
            "image_id": item["image_id"],
            "media_path": media_path or "",
            "media_content_type": media_content_type(media_path or ""),
            "attachment_type": "image",
            "mapping_status": "mapped" if media_path else "media_missing",
            "image_path": "",
        }
        if media_path:
            row["image_path"] = materialize_xlsx_media(
                workbook_path,
                media_path,
                output_dir=output_dir,
                attachment_id=attachment_id,
            )
            if not row["image_path"]:
                row["mapping_status"] = "extract_failed"
        rows.append(row)
    return rows


def materialize_xlsx_media(workbook_path: Path, media_path: str, *, output_dir: Path, attachment_id: str) -> str:
    normalized = normalize_xlsx_target(media_path)
    suffix = Path(normalized).suffix.lower()
    if suffix not in SUPPORTED_IMAGE_SUFFIXES:
        return ""
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            data = archive.read(normalized)
    except (OSError, KeyError, zipfile.BadZipFile):
        return ""
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / f"{safe_key_part(attachment_id)}{suffix}"
    target.write_bytes(data)
    return str(target.resolve())


def registry_by_attachment_id(rows: list[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for row in rows:
        attachment_id = str(row.get("attachment_id") or "")
        if attachment_id and attachment_id not in output:
            output[attachment_id] = dict(row)
    return output


def registry_by_sheet_row(rows: list[Mapping[str, Any]]) -> dict[tuple[str, int], list[dict[str, Any]]]:
    output: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        sheet_name = str(row.get("sheet_name") or "")
        try:
            row_index = int(row.get("row_index") or 0)
        except (TypeError, ValueError):
            continue
        if sheet_name and row_index:
            output[(sheet_name, row_index)].append(dict(row))
    return dict(output)


def attachment_from_registry(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "attachment_id": row.get("attachment_id") or "",
        "file_id": row.get("file_id") or "",
        "sheet_name": row.get("sheet_name") or "",
        "source_cell": row.get("source_cell") or "",
        "media_path": row.get("media_path") or "",
        "media_content_type": row.get("media_content_type") or "",
        "attachment_type": row.get("attachment_type") or "image",
        "image_id": row.get("image_id") or "",
        "image_path": row.get("image_path") or "",
        "source_file_path": row.get("source_file_path") or "",
        "relative_path": row.get("relative_path") or "",
        "mapping_status": row.get("mapping_status") or "",
    }


def enrich_attachment_with_registry(attachment: Mapping[str, Any], registry: Mapping[str, Mapping[str, Any]]) -> dict[str, Any]:
    output = dict(attachment)
    row = registry.get(str(output.get("attachment_id") or ""))
    if not row:
        return output
    for key in (
        "image_path",
        "source_file_path",
        "relative_path",
        "media_path",
        "media_content_type",
        "mapping_status",
        "image_id",
        "file_id",
    ):
        if row.get(key) and not output.get(key):
            output[key] = row[key]
    return output


def relative_path(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return path.name
