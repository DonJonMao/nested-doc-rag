from __future__ import annotations

import argparse
import hashlib
import io
import json
import mimetypes
import posixpath
import re
import shutil
import struct
import subprocess
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import olefile
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook


DEFAULT_PROBED_MANIFEST = Path("/Users/mao/projects/datacenter/artifacts/03_format_probe/probed_manifest.jsonl")
DEFAULT_STRUCTURE_DIR = Path("/Users/mao/projects/datacenter/artifacts/04a_structure_parse")
DEFAULT_OUT_DIR = Path("/Users/mao/projects/datacenter/artifacts/04b_embedded_object_parse")

CELL_RE = re.compile(r"^([A-Z]+)([0-9]+)$")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def split_cell_ref(ref: str | None) -> tuple[int, str] | None:
    match = CELL_RE.match(ref or "")
    if not match:
        return None
    col, row = match.groups()
    return int(row), col


def safe_name(value: str | None, fallback: str) -> str:
    text = value or fallback
    text = text.replace("\\", "_").replace("/", "_").replace(":", "_")
    text = re.sub(r"[\x00-\x1f]+", "_", text).strip(" ._")
    return text or fallback


def sha10(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()[:10]


def display_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def compact_repeated_values(values: list[str]) -> list[str]:
    compacted: list[str] = []
    for value in values:
        if not value:
            continue
        if compacted and compacted[-1] == value:
            continue
        compacted.append(value)
    return compacted


def is_heading_style(style_name: str | None) -> bool:
    if not style_name:
        return False
    return style_name.startswith("Heading") or style_name.startswith("标题")


def heading_level(style_name: str | None) -> int:
    if not style_name:
        return 1
    match = re.search(r"(\d+)", style_name)
    return int(match.group(1)) if match else 1


def update_heading_stack(stack: list[str], text: str, style_name: str | None) -> list[str]:
    level = max(1, heading_level(style_name))
    next_stack = stack[: level - 1]
    next_stack.append(text)
    return next_stack


def contextual_text(text: str, section_context: list[str] | None = None) -> str:
    section_context = section_context or []
    if not section_context:
        return text
    return f"章节：{' / '.join(section_context)}。{text}"


def table_row_text(values: list[str], header: list[str], section_context: list[str]) -> str:
    values = compact_repeated_values(values)
    header = compact_repeated_values(header)
    if not values:
        return ""
    if header and values != header:
        pairs: list[str] = []
        for index, value in enumerate(values):
            key = header[index] if index < len(header) and header[index] else f"列{index + 1}"
            pairs.append(f"{key}：{value}")
        return contextual_text("；".join(pairs), section_context)
    return contextual_text("表头：" + " | ".join(values), section_context)


def table_header_score(values: list[str]) -> int:
    header_keywords = {
        "类别",
        "项目",
        "结果",
        "异常描述",
        "业务链路",
        "接入设备",
        "接入端口",
        "实际带宽",
        "ODF信息",
        "设备名称",
        "设备位置",
        "设备IP",
    }
    return sum(1 for value in values if value in header_keywords)


def choose_table_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, values in enumerate(rows[:3]):
        score = table_header_score(values)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def metadata_row_text(values: list[str], section_context: list[str]) -> str:
    pairs: list[str] = []
    index = 0
    while index < len(values):
        key = values[index]
        value = values[index + 1] if index + 1 < len(values) else ""
        pairs.append(f"{key}：{value}" if value else key)
        index += 2
    return contextual_text("；".join(pairs), section_context)


def decode_bytes(value: bytes) -> str:
    for encoding in ("gb18030", "utf-8", "cp437", "latin1"):
        try:
            return value.decode(encoding).rstrip("\x00")
        except UnicodeDecodeError:
            continue
    return value.decode("latin1", errors="replace").rstrip("\x00")


def read_c_string(data: bytes, pos: int) -> tuple[str, int]:
    end = data.index(b"\x00", pos)
    return decode_bytes(data[pos:end]), end + 1


def repair_zip_name(name: str) -> str:
    try:
        repaired = name.encode("cp437").decode("gb18030")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name
    original_cjk = sum("\u4e00" <= ch <= "\u9fff" for ch in name)
    repaired_cjk = sum("\u4e00" <= ch <= "\u9fff" for ch in repaired)
    if repaired_cjk > original_cjk:
        return repaired
    return name


def media_type_for_path(path: str) -> str | None:
    guessed, _ = mimetypes.guess_type(path)
    return guessed


def detect_file_type(data: bytes, name: str | None = None) -> str:
    lower_name = (name or "").lower()
    if data.startswith(b"%PDF-") or lower_name.endswith(".pdf"):
        return "pdf"
    if data.startswith(b"Rar!\x1a\x07") or lower_name.endswith(".rar"):
        return "rar"
    if data.startswith(b"AC10") or lower_name.endswith(".dwg"):
        return "dwg"
    if data.startswith(b"\x89PNG\r\n\x1a\n") or lower_name.endswith(".png"):
        return "image"
    if data.startswith(b"\xff\xd8\xff") or lower_name.endswith((".jpg", ".jpeg")):
        return "image"
    if lower_name.endswith(".doc"):
        return "doc"
    if lower_name.endswith(".docx"):
        return "docx"
    if lower_name.endswith(".xlsx"):
        return "xlsx"
    if lower_name.endswith(".pptx"):
        return "pptx"
    if zipfile.is_zipfile(io.BytesIO(data)):
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = set(zf.namelist())
            if "word/document.xml" in names:
                return "docx"
            if "xl/workbook.xml" in names:
                return "xlsx"
            if "ppt/presentation.xml" in names:
                return "pptx"
            if "[Content_Types].xml" in names and any(name.startswith("visio/") for name in names):
                return "vsdx"
        return "zip"
    if data and olefile.isOleFile(io.BytesIO(data)):
        try:
            ole = olefile.OleFileIO(io.BytesIO(data))
            streams = {"/".join(item) for item in ole.listdir(streams=True, storages=False)}
            ole.close()
            if "WordDocument" in streams:
                return "doc"
        except Exception:
            pass
    if lower_name.endswith(".vsdx"):
        return "vsdx"
    return "unknown"


def parse_ole10_native(data: bytes) -> dict[str, Any]:
    declared_size = struct.unpack_from("<I", data, 0)[0] if len(data) >= 4 else None
    pos = 4
    flags = struct.unpack_from("<H", data, pos)[0] if len(data) >= pos + 2 else None
    pos += 2
    label, pos = read_c_string(data, pos)
    source_path, pos = read_c_string(data, pos)

    unknown = data[pos : pos + 4].hex() if len(data) >= pos + 4 else None
    pos += 4
    temp_path = ""
    if len(data) >= pos + 4:
        temp_len = struct.unpack_from("<I", data, pos)[0]
        pos += 4
        if 0 <= temp_len <= 4096 and len(data) >= pos + temp_len:
            temp_path = decode_bytes(data[pos : pos + temp_len])
            pos += temp_len
    native_size = struct.unpack_from("<I", data, pos)[0] if len(data) >= pos + 4 else 0
    pos += 4
    if native_size <= 0 or native_size > len(data) - pos:
        native_size = len(data) - pos
    payload = data[pos : pos + native_size]
    return {
        "native_declared_size": declared_size,
        "native_flags": flags,
        "embedded_file_name": label,
        "embedded_source_path": source_path,
        "embedded_temp_path": temp_path,
        "native_unknown": unknown,
        "payload_size": len(payload),
        "payload": payload,
    }


def extract_payload_from_ole(ole_bytes: bytes, attachment: dict[str, Any]) -> dict[str, Any]:
    if not olefile.isOleFile(io.BytesIO(ole_bytes)):
        embedded_name = posixpath.basename(attachment.get("media_path") or "") or attachment.get("prog_id") or "embedded_object"
        return {
            "embedded_file_name": embedded_name,
            "embedded_source_path": None,
            "embedded_temp_path": None,
            "payload_size": len(ole_bytes),
            "ole_streams": [],
            "payload": ole_bytes,
            "payload_source": "direct_embedding",
        }
    ole = olefile.OleFileIO(io.BytesIO(ole_bytes))
    streams = ["/".join(item) for item in ole.listdir(streams=True, storages=False)]
    try:
        if ole.exists("\x01Ole10Native"):
            native = parse_ole10_native(ole.openstream("\x01Ole10Native").read())
            payload = native.pop("payload")
            return {
                **native,
                "ole_streams": streams,
                "payload": payload,
                "payload_source": "Ole10Native",
            }
        if ole.exists("package"):
            payload = ole.openstream("package").read()
            return {
                "embedded_file_name": None,
                "embedded_source_path": None,
                "embedded_temp_path": None,
                "payload_size": len(payload),
                "ole_streams": streams,
                "payload": payload,
                "payload_source": "package",
            }
        if ole.exists("WordDocument"):
            return {
                "embedded_file_name": f"{attachment.get('prog_id') or 'Word.Document.8'}.doc",
                "embedded_source_path": None,
                "embedded_temp_path": None,
                "payload_size": len(ole_bytes),
                "ole_streams": streams,
                "payload": ole_bytes,
                "payload_source": "ole_word_document",
            }
        return {
            "embedded_file_name": None,
            "embedded_source_path": None,
            "embedded_temp_path": None,
            "payload_size": 0,
            "ole_streams": streams,
            "payload": b"",
            "payload_source": "none",
            "parse_error": "no Ole10Native or package stream",
        }
    finally:
        ole.close()


def parent_segment_id(attachment: dict[str, Any], sheet_index_by_name: dict[str, int]) -> str | None:
    parsed = split_cell_ref(attachment.get("source_cell"))
    if not parsed:
        return None
    row, _ = parsed
    sheet_index = sheet_index_by_name.get(attachment.get("sheet_name", ""))
    if sheet_index is None:
        return None
    return f"seg_{attachment['file_id']}_{sheet_index:02d}_row_{row:04d}"


def source_chain(object_record: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "level": "knowledge_base_row",
            "segment_id": object_record.get("parent_segment_id"),
            "file_name": object_record.get("parent_file_name"),
            "sheet_name": object_record.get("parent_sheet_name"),
            "source_cell": object_record.get("parent_source_cell"),
        },
        {
            "level": "embedded_object",
            "attachment_id": object_record["parent_attachment_id"],
            "file_name": object_record.get("embedded_file_name"),
            "file_type": object_record.get("embedded_file_type"),
        },
    ]


def make_segment(
    object_record: dict[str, Any],
    segment_type: str,
    ordinal: int,
    text: str,
    local_anchor: dict[str, Any],
) -> dict[str, Any]:
    text = display_text(text)
    section_context = local_anchor.get("section_context") or []
    table_header = local_anchor.get("table_header") or []
    segment_id = f"embseg_{object_record['parent_attachment_id']}_{ordinal:04d}"
    embedding_parts = [
        f"数据中心：{object_record.get('data_center_id') or 'global'}。",
        f"父文件：{object_record.get('parent_file_name')}。",
        f"父位置：{object_record.get('parent_sheet_name')}!{object_record.get('parent_source_cell')}。",
        f"嵌入文件：{object_record.get('embedded_file_name') or object_record.get('embedded_payload_path')}。",
        f"章节：{' / '.join(section_context)}。" if section_context else "",
        f"表头：{' | '.join(table_header)}。" if table_header else "",
        f"内容：{text}。",
    ]
    return {
        "segment_id": segment_id,
        "segment_type": segment_type,
        "parent_segment_id": object_record.get("parent_segment_id"),
        "parent_attachment_id": object_record["parent_attachment_id"],
        "parent_file_id": object_record["parent_file_id"],
        "parent_file_name": object_record["parent_file_name"],
        "parent_sheet_name": object_record.get("parent_sheet_name"),
        "parent_source_cell": object_record.get("parent_source_cell"),
        "data_center_id": object_record.get("data_center_id") or "global",
        "embedded_object_id": object_record["embedded_object_id"],
        "embedded_file_name": object_record.get("embedded_file_name"),
        "embedded_file_type": object_record.get("embedded_file_type"),
        "local_anchor": local_anchor,
        "source_chain": source_chain(object_record),
        "raw_text": text,
        "embedding_text": "".join(embedding_parts),
    }


def parse_docx_segments(data: bytes, object_record: dict[str, Any], ordinal_start: int = 1) -> list[dict[str, Any]]:
    doc = Document(io.BytesIO(data))
    segments: list[dict[str, Any]] = []
    ordinal = ordinal_start
    paragraph_index = 0
    table_index = 0
    heading_stack: list[str] = []
    for child in doc.element.body.iterchildren():
        if child.tag == qn("w:p"):
            paragraph_index += 1
            paragraph = Paragraph(child, doc)
            text = paragraph.text.strip()
            if not text:
                continue
            style_name = paragraph.style.name if paragraph.style else None
            if is_heading_style(style_name):
                heading_stack = update_heading_stack(heading_stack, text, style_name)
            segments.append(
                make_segment(
                    object_record,
                    "embedded_docx_paragraph",
                    ordinal,
                    contextual_text(text, heading_stack[:-1] if is_heading_style(style_name) else heading_stack),
                    {
                        "block_type": "paragraph",
                        "paragraph_index": paragraph_index,
                        "style_name": style_name,
                        "section_context": heading_stack[:],
                    },
                )
            )
            ordinal += 1
        elif child.tag == qn("w:tbl"):
            table_index += 1
            table = Table(child, doc)
            table_rows = [
                [display_text(cell.text) for cell in row.cells]
                for row in table.rows
            ]
            table_rows = [values for values in table_rows if any(values)]
            if not table_rows:
                continue
            header_row_index = choose_table_header_row(table_rows)
            header_values = table_rows[header_row_index]
            for row_index, values in enumerate(table_rows, start=1):
                if row_index - 1 < header_row_index:
                    text = metadata_row_text(values, heading_stack)
                else:
                    text = table_row_text(values, header_values, heading_stack)
                if not text:
                    continue
                segments.append(
                    make_segment(
                        object_record,
                        "embedded_docx_table_row",
                        ordinal,
                        text,
                        {
                            "block_type": "table_row",
                            "table_index": table_index,
                            "row_index": row_index,
                            "section_context": heading_stack[:],
                            "table_header": header_values[:] if row_index - 1 >= header_row_index else [],
                            "row_values": values[:],
                            "table_header_row_index": header_row_index + 1,
                        },
                    )
                )
                ordinal += 1
    return segments


def parse_xlsx_segments(data: bytes, object_record: dict[str, Any], ordinal_start: int = 1) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    segments: list[dict[str, Any]] = []
    ordinal = ordinal_start
    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [display_text(value) for value in row if display_text(value)]
            if not values:
                continue
            text = " | ".join(values)
            segments.append(
                make_segment(
                    object_record,
                    "embedded_xlsx_row",
                    ordinal,
                    text,
                    {"sheet_index": sheet_index, "sheet_name": worksheet.title, "row_index": row_index},
                )
            )
            ordinal += 1
    workbook.close()
    return segments


def parse_plain_text_segments(
    text: str,
    object_record: dict[str, Any],
    segment_type: str,
    anchor_key: str,
    ordinal_start: int = 1,
) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    ordinal = ordinal_start
    blocks = [display_text(block) for block in re.split(r"\n\s*\n+", text) if display_text(block)]
    for index, block in enumerate(blocks, start=1):
        segments.append(
            make_segment(
                object_record,
                segment_type,
                ordinal,
                block,
                {"block_type": anchor_key, f"{anchor_key}_index": index},
            )
        )
        ordinal += 1
    return segments


def parse_pdf_segments(data: bytes, object_record: dict[str, Any], ordinal_start: int = 1) -> tuple[list[dict[str, Any]], list[str]]:
    diagnostics: list[str] = []
    try:
        from pypdf import PdfReader
    except Exception as exc:
        return [], [f"pypdf unavailable: {exc!r}"]

    try:
        reader = PdfReader(io.BytesIO(data))
        segments: list[dict[str, Any]] = []
        ordinal = ordinal_start
        for page_index, page in enumerate(reader.pages, start=1):
            text = display_text(page.extract_text() or "")
            if not text:
                continue
            segments.append(
                make_segment(
                    object_record,
                    "embedded_pdf_page",
                    ordinal,
                    text,
                    {"block_type": "pdf_page", "page_index": page_index},
                )
            )
            ordinal += 1
        if not segments:
            diagnostics.append("pdf has no extractable text; keep as proof asset, no OCR")
        return segments, diagnostics
    except Exception as exc:
        return [], [repr(exc)]


def parse_doc_segments(data: bytes, object_record: dict[str, Any], payload_path: Path, work_dir: Path, ordinal_start: int = 1) -> tuple[list[dict[str, Any]], list[str]]:
    textutil = shutil.which("textutil")
    if not textutil:
        return [], ["textutil not available for legacy Word document"]
    doc_dir = work_dir / "legacy_doc"
    doc_dir.mkdir(parents=True, exist_ok=True)
    doc_path = doc_dir / f"{payload_path.stem or 'legacy_word'}.doc"
    txt_path = doc_dir / f"{payload_path.stem or 'legacy_word'}.txt"
    doc_path.write_bytes(data)
    proc = subprocess.run(
        [textutil, "-convert", "txt", "-output", str(txt_path), str(doc_path)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    if proc.returncode != 0 or not txt_path.exists():
        return [], [proc.stdout[-2000:] or "textutil failed"]
    text = txt_path.read_text(encoding="utf-8", errors="replace")
    segments = parse_plain_text_segments(text, object_record, "embedded_doc_paragraph", "paragraph", ordinal_start)
    return segments, [proc.stdout[-2000:]] if proc.stdout.strip() else []


def xml_text_values(data: bytes, wanted_tags: set[str] | None = None) -> list[str]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    values: list[str] = []
    for node in root.iter():
        node_name = node.tag.rsplit("}", 1)[-1] if "}" in node.tag else node.tag
        if wanted_tags is not None and node_name not in wanted_tags:
            continue
        if node.text and display_text(node.text):
            values.append(display_text(node.text))
    return values


def parse_vsdx_segments(data: bytes, object_record: dict[str, Any], ordinal_start: int = 1) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    ordinal = ordinal_start
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        page_names = sorted(name for name in zf.namelist() if name.startswith("visio/pages/") and name.endswith(".xml"))
        for page_index, name in enumerate(page_names, start=1):
            values = xml_text_values(zf.read(name))
            text = display_text("；".join(values))
            if not text:
                continue
            segments.append(
                make_segment(
                    object_record,
                    "embedded_vsdx_page",
                    ordinal,
                    text,
                    {"block_type": "vsdx_page", "page_index": page_index, "entry_name": name},
                )
            )
            ordinal += 1
    return segments


def parse_pptx_segments(data: bytes, object_record: dict[str, Any], ordinal_start: int = 1) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    ordinal = ordinal_start
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        slide_names = sorted(name for name in zf.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        for slide_index, name in enumerate(slide_names, start=1):
            values = xml_text_values(zf.read(name), {"t"})
            text = display_text("；".join(values))
            if not text:
                continue
            segments.append(
                make_segment(
                    object_record,
                    "embedded_pptx_slide",
                    ordinal,
                    text,
                    {"block_type": "pptx_slide", "slide_index": slide_index, "entry_name": name},
                )
            )
            ordinal += 1
    return segments


def parse_known_child_segments(
    child_type: str,
    child_data: bytes,
    child_object: dict[str, Any],
    child_path: Path,
    work_dir: Path,
    ordinal_start: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    if child_type == "docx":
        return parse_docx_segments(child_data, child_object, ordinal_start), []
    if child_type == "xlsx":
        return parse_xlsx_segments(child_data, child_object, ordinal_start), []
    if child_type == "pdf":
        return parse_pdf_segments(child_data, child_object, ordinal_start)
    if child_type == "doc":
        return parse_doc_segments(child_data, child_object, child_path, work_dir, ordinal_start)
    if child_type == "vsdx":
        return parse_vsdx_segments(child_data, child_object, ordinal_start), []
    if child_type == "pptx":
        return parse_pptx_segments(child_data, child_object, ordinal_start), []
    return [], []


def extract_archive_with_7z(payload_path: Path, out_dir: Path) -> tuple[bool, str]:
    seven_zip = shutil.which("7z") or shutil.which("7zz")
    if not seven_zip:
        return False, "7z not available"
    out_dir.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [seven_zip, "x", "-y", f"-o{out_dir}", str(payload_path)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    return proc.returncode == 0, proc.stdout[-2000:]


def extract_archive_with_bsdtar(payload_path: Path, out_dir: Path) -> tuple[bool, str]:
    bsdtar = shutil.which("bsdtar")
    if not bsdtar:
        return False, "bsdtar not available"
    out_dir.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [bsdtar, "-xf", str(payload_path), "-C", str(out_dir)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    return proc.returncode == 0, proc.stdout[-2000:]


def extract_archive(payload_path: Path, out_dir: Path) -> tuple[bool, list[str]]:
    diagnostics: list[str] = []
    ok, log = extract_archive_with_7z(payload_path, out_dir)
    diagnostics.append(log)
    if ok:
        return True, diagnostics
    ok, log = extract_archive_with_bsdtar(payload_path, out_dir)
    diagnostics.append(log)
    return ok, diagnostics


def list_archive_with_7z(payload_path: Path) -> tuple[list[dict[str, Any]], str]:
    seven_zip = shutil.which("7z") or shutil.which("7zz")
    if not seven_zip:
        return [], "7z not available"
    proc = subprocess.run(
        [seven_zip, "l", "-slt", str(payload_path)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    entries: list[dict[str, Any]] = []
    current: dict[str, str] = {}

    def flush() -> None:
        nonlocal current
        if current:
            entry = archive_entry_from_7z(current, archive_path=str(payload_path))
            if entry:
                entries.append(entry)
            current = {}

    for line in proc.stdout.splitlines():
        if " = " not in line:
            continue
        key, value = line.split(" = ", 1)
        if key == "Path" and current:
            flush()
        if key in {"Path", "Folder", "Size", "Packed Size", "Modified", "Encrypted", "Method"}:
            current[key] = value
    flush()
    return entries, proc.stdout[-2000:]


def archive_entry_from_7z(values: dict[str, str], archive_path: str) -> dict[str, Any] | None:
    name = values.get("Path", "")
    if not name or name == archive_path or values.get("Folder") == "+":
        return None
    try:
        size = int(values.get("Size", "0"))
    except ValueError:
        size = 0
    return {
        "name": name,
        "size": size,
        "file_type": detect_file_type(b"", name),
        "path": None,
        "archive_method": values.get("Method"),
        "encrypted": values.get("Encrypted"),
        "modified": values.get("Modified"),
    }


def parse_payload_segments(
    data: bytes,
    object_record: dict[str, Any],
    payload_path: Path,
    work_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, list[str]]:
    file_type = object_record["embedded_file_type"]
    segments: list[dict[str, Any]] = []
    children: list[dict[str, Any]] = []
    diagnostics: list[str] = []

    try:
        if file_type == "docx":
            segments.extend(parse_docx_segments(data, object_record))
            return segments, children, "parsed", diagnostics
        if file_type == "xlsx":
            segments.extend(parse_xlsx_segments(data, object_record))
            return segments, children, "parsed", diagnostics
        if file_type == "pdf":
            parsed_segments, logs = parse_pdf_segments(data, object_record)
            segments.extend(parsed_segments)
            diagnostics.extend(logs)
            return segments, children, "parsed" if parsed_segments else "extracted_only", diagnostics
        if file_type == "doc":
            parsed_segments, logs = parse_doc_segments(data, object_record, payload_path, work_dir)
            segments.extend(parsed_segments)
            diagnostics.extend(logs)
            return segments, children, "parsed" if parsed_segments else "extracted_only", diagnostics
        if file_type == "vsdx":
            segments.extend(parse_vsdx_segments(data, object_record))
            return segments, children, "parsed" if segments else "extracted_only", diagnostics
        if file_type == "pptx":
            segments.extend(parse_pptx_segments(data, object_record))
            return segments, children, "parsed" if segments else "extracted_only", diagnostics
        if file_type in {"zip", "vsdx"}:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                child_dir = work_dir / "zip_entries"
                child_dir.mkdir(parents=True, exist_ok=True)
                ordinal = 1
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    child_name = repair_zip_name(info.filename)
                    child_data = zf.read(info)
                    child_type = detect_file_type(child_data, child_name)
                    child_path = child_dir / f"{ordinal:03d}_{safe_name(posixpath.basename(child_name), 'entry')}"
                    child_path.write_bytes(child_data)
                    child_record = {
                        "name": child_name,
                        "size": len(child_data),
                        "file_type": child_type,
                        "path": str(child_path),
                    }
                    children.append(child_record)
                    child_object = {**object_record, "embedded_file_name": child_name, "embedded_file_type": child_type}
                    parsed_segments, logs = parse_known_child_segments(
                        child_type,
                        child_data,
                        child_object,
                        child_path,
                        child_dir,
                        len(segments) + 1,
                    )
                    segments.extend(parsed_segments)
                    diagnostics.extend(logs)
                    ordinal += 1
            return segments, children, "parsed_archive", diagnostics
        if file_type == "rar":
            extract_dir = work_dir / "rar_entries"
            ok, logs = extract_archive(payload_path, extract_dir)
            diagnostics.extend(logs)
            if not ok:
                listed_children, list_log = list_archive_with_7z(payload_path)
                diagnostics.append(list_log)
                children.extend(listed_children)
                return segments, children, "listed_archive" if listed_children else "unsupported_archive", diagnostics
            ordinal = 1
            for child_path in sorted(path for path in extract_dir.rglob("*") if path.is_file()):
                child_data = child_path.read_bytes()
                rel_name = str(child_path.relative_to(extract_dir))
                child_type = detect_file_type(child_data, rel_name)
                children.append({"name": rel_name, "size": len(child_data), "file_type": child_type, "path": str(child_path)})
                child_object = {**object_record, "embedded_file_name": rel_name, "embedded_file_type": child_type}
                parsed_segments, logs = parse_known_child_segments(
                    child_type,
                    child_data,
                    child_object,
                    child_path,
                    extract_dir,
                    len(segments) + 1,
                )
                segments.extend(parsed_segments)
                diagnostics.extend(logs)
                ordinal += 1
            return segments, children, "parsed_archive", diagnostics
    except Exception as exc:
        diagnostics.append(repr(exc))
        return segments, children, "error", diagnostics

    return segments, children, "extracted_only", diagnostics


def sheet_index_maps(structure_dir: Path, file_id: str) -> dict[str, int]:
    workbook_path = structure_dir / "workbooks" / f"{file_id}.workbook.json"
    if not workbook_path.exists():
        return {}
    workbook = read_json(workbook_path)
    return {sheet["sheet_name"]: sheet["sheet_index"] for sheet in workbook.get("sheets", [])}


def run(probed_manifest: Path, structure_dir: Path, out_dir: Path) -> dict[str, Any]:
    if out_dir.exists():
        shutil.rmtree(out_dir)
    files_dir = out_dir / "embedded_files"
    files_dir.mkdir(parents=True, exist_ok=True)

    source_records = {record["file_id"]: record for record in read_jsonl(probed_manifest)}
    attachments: list[dict[str, Any]] = []
    for path in sorted((structure_dir / "attachments").glob("*.attachments.jsonl")):
        attachments.extend(item for item in read_jsonl(path) if item.get("attachment_type") == "embedded_object")

    objects: list[dict[str, Any]] = []
    segments: list[dict[str, Any]] = []
    for attachment in attachments:
        source = source_records.get(attachment["file_id"])
        if not source:
            continue
        sheet_index_by_name = sheet_index_maps(structure_dir, attachment["file_id"])
        object_dir = files_dir / attachment["attachment_id"]
        object_dir.mkdir(parents=True, exist_ok=True)
        try:
            with zipfile.ZipFile(source["source_path"]) as zf:
                ole_bytes = zf.read(attachment["media_path"])
        except Exception as exc:
            objects.append(
                {
                    "embedded_object_id": f"embobj_{attachment['attachment_id']}",
                    "parent_attachment_id": attachment["attachment_id"],
                    "parent_file_id": attachment["file_id"],
                    "parse_status": "error",
                    "diagnostics": [repr(exc)],
                }
            )
            continue

        ole_path = object_dir / "ole_object.bin"
        ole_path.write_bytes(ole_bytes)
        payload_info = extract_payload_from_ole(ole_bytes, attachment)
        payload = payload_info.pop("payload")
        embedded_name = payload_info.get("embedded_file_name") or attachment.get("prog_id") or "embedded_object"
        file_type = detect_file_type(payload, embedded_name)
        suffix = Path(embedded_name).suffix if Path(embedded_name).suffix else f".{file_type}" if file_type != "unknown" else ".bin"
        payload_name = safe_name(embedded_name, f"payload_{sha10(payload)}{suffix}")
        payload_path = object_dir / payload_name
        if payload:
            payload_path.write_bytes(payload)

        parent_id = parent_segment_id(attachment, sheet_index_by_name)
        object_record = {
            "embedded_object_id": f"embobj_{attachment['attachment_id']}",
            "parent_attachment_id": attachment["attachment_id"],
            "parent_segment_id": parent_id,
            "parent_file_id": attachment["file_id"],
            "parent_file_name": source["file_name"],
            "parent_relative_path": source["relative_path"],
            "parent_sheet_name": attachment.get("sheet_name"),
            "parent_source_cell": attachment.get("source_cell"),
            "data_center_id": source.get("data_center_id") or "global",
            "document_role": source.get("document_role"),
            "prog_id": attachment.get("prog_id"),
            "object_shape_id": attachment.get("object_shape_id"),
            "ole_media_path": attachment.get("media_path"),
            "ole_file_path": str(ole_path),
            "embedded_file_name": embedded_name,
            "embedded_file_type": file_type,
            "embedded_payload_path": str(payload_path) if payload else None,
            "embedded_payload_size": len(payload),
            "preview_media_path": attachment.get("preview_media_path"),
            **{key: value for key, value in payload_info.items() if key != "parse_error"},
        }
        child_segments, children, parse_status, diagnostics = parse_payload_segments(payload, object_record, payload_path, object_dir)
        object_record["parse_status"] = parse_status
        object_record["child_file_count"] = len(children)
        object_record["child_segment_count"] = len(child_segments)
        object_record["children"] = children[:50]
        object_record["diagnostics"] = diagnostics
        if payload_info.get("parse_error"):
            object_record["diagnostics"].append(payload_info["parse_error"])
        objects.append(object_record)
        segments.extend(child_segments)

    summary = {
        "total_embedded_objects": len(objects),
        "total_embedded_segments": len(segments),
        "object_counts_by_type": dict(Counter(item.get("embedded_file_type") for item in objects)),
        "parse_status_counts": dict(Counter(item.get("parse_status") for item in objects)),
        "segment_counts_by_type": dict(Counter(item.get("segment_type") for item in segments)),
    }
    write_jsonl(out_dir / "embedded_objects.jsonl", objects)
    write_jsonl(out_dir / "embedded_segments.jsonl", segments)
    write_json(out_dir / "summary.json", summary)
    write_visualization(out_dir, summary, objects, segments)
    return summary


def write_visualization(out_dir: Path, summary: dict[str, Any], objects: list[dict[str, Any]], segments: list[dict[str, Any]]) -> None:
    lines: list[str] = []
    lines.append("# Step 04B 嵌入对象下钻解析可视化\n")
    lines.append(f"- 嵌入对象数：**{summary['total_embedded_objects']}**")
    lines.append(f"- 可向量化子 segment 数：**{summary['total_embedded_segments']}**")
    lines.append("- 本步骤只解析可读文本容器；图片不 OCR，PDF/RAR 中不可解析内容只保留文件和父标签。\n")

    lines.append("## 类型统计\n")
    lines.append("| embedded_file_type | count |")
    lines.append("|---|---:|")
    for key, count in sorted(summary["object_counts_by_type"].items()):
        lines.append(f"| `{key}` | {count} |")

    lines.append("\n## 解析状态\n")
    lines.append("| parse_status | count |")
    lines.append("|---|---:|")
    for key, count in sorted(summary["parse_status_counts"].items()):
        lines.append(f"| `{key}` | {count} |")

    lines.append("\n## 对象样例\n")
    lines.append("| parent | cell | embedded file | type | status | child files | child segments |")
    lines.append("|---|---|---|---|---|---:|---:|")
    for item in objects[:20]:
        lines.append(
            f"| `{item.get('parent_file_name')}` | `{item.get('parent_source_cell')}` | "
            f"`{item.get('embedded_file_name')}` | `{item.get('embedded_file_type')}` | "
            f"`{item.get('parse_status')}` | {item.get('child_file_count', 0)} | {item.get('child_segment_count', 0)} |"
        )

    lines.append("\n## 子 Segment 样例\n")
    lines.append("| parent_cell | embedded_file | segment_type | text |")
    lines.append("|---|---|---|---|")
    for segment in segments[:20]:
        text = display_text(segment.get("raw_text", ""))
        if len(text) > 100:
            text = text[:99] + "..."
        lines.append(
            f"| `{segment.get('parent_source_cell')}` | `{segment.get('embedded_file_name')}` | "
            f"`{segment.get('segment_type')}` | {text} |"
        )

    (out_dir / "visualization.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Drill into embedded OLE objects and extract parseable child segments.")
    parser.add_argument("--probed-manifest", type=Path, default=DEFAULT_PROBED_MANIFEST)
    parser.add_argument("--structure-dir", type=Path, default=DEFAULT_STRUCTURE_DIR)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()
    summary = run(args.probed_manifest, args.structure_dir, args.out_dir)
    print(
        f"parsed {summary['total_embedded_objects']} embedded objects and "
        f"{summary['total_embedded_segments']} child segments -> {args.out_dir}"
    )


if __name__ == "__main__":
    main()
